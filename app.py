import calendar
import hmac
import os
import re
import secrets
import shutil
import smtplib
import sqlite3
import tempfile
import threading
import time as _time
import uuid
from io import BytesIO
from decimal import Decimal
from datetime import date, datetime, time, timedelta, timezone
from zoneinfo import ZoneInfo
from email.message import EmailMessage
from functools import wraps
from pathlib import Path
from typing import Optional

from flask import (
    Flask,
    after_this_request,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_sqlalchemy import SQLAlchemy
from flask_wtf.csrf import CSRFProtect
from openpyxl import Workbook
from sqlalchemy import delete as sa_delete
from sqlalchemy import exists, inspect as sa_inspect, select, text as sa_text
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy.orm import joinedload
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from PIL import Image, UnidentifiedImageError

import requests


app = Flask(__name__)
# Persist SQLite DB in /app/instance (volume-mounted in compose).
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URL", "sqlite:////app/instance/emtk.db")


def _migrate_legacy_sqlite_filename(database_uri: str) -> None:
    """
    Backward compatibility: previously the SQLite file was named ``smart_zhk.db``.
    If the configured URI points to a file-based SQLite DB and the new file does
    not exist yet, but a legacy ``smart_zhk.db`` is present in the same folder,
    rename it so running deployments keep their data after the rebuild.
    """
    if not database_uri.startswith("sqlite:"):
        return
    try:
        # SQLAlchemy URL forms: sqlite:///relative/path.db  or  sqlite:////abs/path.db
        path_part = database_uri.split("sqlite:", 1)[1].lstrip("/")
        if not path_part or path_part == ":memory:":
            return
        new_path = Path("/" + path_part) if database_uri.startswith("sqlite:////") else Path(path_part)
        legacy_path = new_path.with_name("smart_zhk.db")
        if legacy_path.exists() and not new_path.exists():
            new_path.parent.mkdir(parents=True, exist_ok=True)
            legacy_path.rename(new_path)
    except OSError:
        # Do not crash startup because of a rename issue; SQLAlchemy will surface
        # a clearer error if the DB is truly unreachable.
        pass


_migrate_legacy_sqlite_filename(app.config["SQLALCHEMY_DATABASE_URI"])
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
debug_mode = os.getenv("FLASK_DEBUG", "0") == "1"
secret_key = os.getenv("SECRET_KEY")
if not secret_key:
    if debug_mode:
        # Development-only fallback to avoid shipping a static weak key.
        secret_key = "dev-only-secret-change-me"
    else:
        raise RuntimeError("SECRET_KEY must be set in production.")
app.config["SECRET_KEY"] = secret_key
app.config["UPLOAD_FOLDER"] = os.path.join("static", "uploads")

# Ограничение размера тела запроса: 5 MB для обычных загрузок (картинки worklog и т.п.).
# Для /admin/settings/database-import лимит поднимается до DB_IMPORT_MAX_BYTES
# через before_request-хук (см. _adjust_max_content_length).
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
DB_IMPORT_MAX_BYTES = 100 * 1024 * 1024
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_BYTES


@app.before_request
def _adjust_max_content_length():
    """Для импорта БД разрешаем большие файлы, не расширяя глобальный лимит."""
    if request.endpoint == "admin_database_import":
        request.max_content_length = DB_IMPORT_MAX_BYTES


@app.errorhandler(413)
def _handle_payload_too_large(_err):
    limit_mb = (
        DB_IMPORT_MAX_BYTES if request.endpoint == "admin_database_import" else MAX_UPLOAD_BYTES
    ) // (1024 * 1024)
    if request.accept_mimetypes.best == "application/json":
        return {"ok": False, "error": f"Fayl ölçüsü {limit_mb} MB-dan böyükdür."}, 413
    flash(f"Fayl ölçüsü {limit_mb} MB-dan böyükdür.", "danger")
    # Redirect только на свой хост, чтобы не превратить 413-хендлер в open-redirect.
    target = url_for("dashboard")
    referer = request.referrer or ""
    if referer and request.host_url and referer.startswith(request.host_url):
        target = referer
    return redirect(target)

# Session cookie hardening (behind HTTPS reverse proxy).
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = os.getenv("SESSION_COOKIE_SAMESITE", "Lax")
app.config["SESSION_COOKIE_SECURE"] = os.getenv("SESSION_COOKIE_SECURE", "1") == "1"

# За reverse proxy (nginx/caddy/traefik) Flask сам видит http/локальный хост,
# поэтому url_for(_external=True) генерировал бы http-ссылки. ProxyFix
# учитывает X-Forwarded-Proto/Host/For. Число прокси регулируется TRUSTED_PROXIES.
try:
    from werkzeug.middleware.proxy_fix import ProxyFix
    _trusted = max(1, int(os.getenv("TRUSTED_PROXIES", "1")))
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=_trusted, x_proto=_trusted, x_host=_trusted, x_prefix=_trusted)
except Exception:
    pass

# Предпочтительная схема для url_for(_external=True) если прокси не выставил заголовок.
app.config["PREFERRED_URL_SCHEME"] = os.getenv("PREFERRED_URL_SCHEME", "https")

# CSRF protection for all mutating requests.
csrf = CSRFProtect(app)

# Basic rate limiting (in-memory). For multi-instance production, use Redis storage.
limiter = Limiter(get_remote_address, app=app, default_limits=[])
db = SQLAlchemy(app)


def _app_timezone():
    """IANA timezone for displaying stored UTC timestamps (env TZ, default Asia/Baku)."""
    tz_name = (os.getenv("TZ") or "Asia/Baku").strip()
    try:
        return ZoneInfo(tz_name)
    except Exception:
        return timezone.utc


def month_sql_expr(column):
    """
    Portable SQL expression that formats a datetime column as 'YYYY-MM'.
    Подбирает функцию под текущий диалект БД:
    - SQLite     -> strftime('%Y-%m', col)
    - PostgreSQL -> to_char(col, 'YYYY-MM')
    - прочие     -> fallback через EXTRACT + LPAD + конкатенация.
    """
    dialect = db.engine.dialect.name
    if dialect == "sqlite":
        return db.func.strftime("%Y-%m", column)
    if dialect in ("postgresql", "postgres"):
        return db.func.to_char(column, "YYYY-MM")
    year = db.func.lpad(db.cast(db.func.extract("year", column), db.String), 4, "0")
    month = db.func.lpad(db.cast(db.func.extract("month", column), db.String), 2, "0")
    return year.op("||")(db.literal("-")).op("||")(month)


def utc_to_local(dt: Optional[datetime]) -> datetime:
    """Convert app UTC datetimes to local wall time (server zone)."""
    if dt is None:
        dt = datetime.now(timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(_app_timezone())


@app.context_processor
def inject_now():
    return {"now": lambda: datetime.now(timezone.utc)}


@app.template_filter("azn")
def azn(value):
    try:
        return f"{float(value):.2f} AZN"
    except (TypeError, ValueError):
        return "0.00 AZN"


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(120), nullable=False)
    phone = db.Column(db.String(30), nullable=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default="resident")
    whatsapp_connected = db.Column(db.Boolean, nullable=False, default=False)
    whatsapp_connected_at = db.Column(db.DateTime, nullable=True)
    whatsapp_jid = db.Column(db.String(64), nullable=True)


class Building(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False, unique=True)
    address = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class Apartment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.String(20), unique=True, nullable=False)
    floor = db.Column(db.Integer, nullable=False)
    rooms = db.Column(db.Integer, nullable=True)
    area = db.Column(db.Float, nullable=False)
    owner_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    owner = db.relationship("User", backref="apartments")
    credit_balance = db.Column(db.Numeric(12, 2), nullable=False, default=Decimal("0.00"))
    building_id = db.Column(db.Integer, db.ForeignKey("building.id"), nullable=True)
    building = db.relationship("Building", backref="apartments")


class Tariff(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    type = db.Column(db.String(20), nullable=False)  # per_m2 | fixed
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    is_active = db.Column(db.Boolean, nullable=False, default=True)


class ApartmentPreset(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    rooms = db.Column(db.Integer, nullable=False)
    area = db.Column(db.Float, nullable=False)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class TariffApartment(db.Model):
    __table_args__ = (db.UniqueConstraint("tariff_id", "apartment_id", name="uq_tariff_apartment"),)
    id = db.Column(db.Integer, primary_key=True)
    tariff_id = db.Column(db.Integer, db.ForeignKey("tariff.id"), nullable=False)
    apartment_id = db.Column(db.Integer, db.ForeignKey("apartment.id"), nullable=False)
    tariff = db.relationship("Tariff", backref="apartment_links")
    apartment = db.relationship("Apartment", backref="tariff_links")


class Invoice(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    apartment_id = db.Column(db.Integer, db.ForeignKey("apartment.id"), nullable=False)
    apartment = db.relationship("Apartment", backref="invoices")
    period = db.Column(db.String(7), nullable=False)  # YYYY-MM
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    paid_amount = db.Column(db.Numeric(12, 2), nullable=False, default=Decimal("0.00"))
    status = db.Column(db.String(20), nullable=False, default="gozlemede")
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


# Invoice.status when fully paid (legacy DB value was ASCII "odenilib").
INVOICE_STATUS_PAID = "ödənilib"


class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    invoice_id = db.Column(db.Integer, db.ForeignKey("invoice.id"), nullable=False)
    invoice = db.relationship("Invoice", backref="payments")
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    comment = db.Column(db.String(255), nullable=True)
    status = db.Column(db.String(20), nullable=False, default="confirmed")
    reviewer_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class SmtpConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    host = db.Column(db.String(120), nullable=True)
    port = db.Column(db.Integer, nullable=False, default=587)
    username = db.Column(db.String(120), nullable=True)
    password = db.Column(db.String(255), nullable=True)
    sender_email = db.Column(db.String(120), nullable=True)
    use_tls = db.Column(db.Boolean, nullable=False, default=True)
    system_name = db.Column(db.String(120), nullable=True)
    house_address = db.Column(db.String(255), nullable=True)
    commandant_name = db.Column(db.String(120), nullable=True)
    contact_phone = db.Column(db.String(50), nullable=True)
    portal_url = db.Column(db.String(255), nullable=True)
    whatsapp_group_url = db.Column(db.String(512), nullable=True)


class WhatsappConfig(db.Model):
    """Evolution API (WhatsApp) integration settings — singleton row."""
    id = db.Column(db.Integer, primary_key=True)
    enabled = db.Column(db.Boolean, nullable=False, default=False)
    api_url = db.Column(db.String(255), nullable=True)
    api_key = db.Column(db.String(255), nullable=True)
    instance = db.Column(db.String(120), nullable=True)
    service_number = db.Column(db.String(30), nullable=True)  # номер бота, куда пишет резидент
    bulk_limit = db.Column(db.Integer, nullable=False, default=10)
    bulk_window_sec = db.Column(db.Integer, nullable=False, default=300)
    webhook_secret = db.Column(db.String(64), nullable=True)


class WhatsappQueue(db.Model):
    """Очередь исходящих WhatsApp-сообщений для rate-limited рассылки."""
    id = db.Column(db.Integer, primary_key=True)
    recipient_phone = db.Column(db.String(30), nullable=False)
    text = db.Column(db.Text, nullable=False)
    invoice_id = db.Column(db.Integer, db.ForeignKey("invoice.id"), nullable=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    status = db.Column(db.String(20), nullable=False, default="pending")  # pending|sent|failed|skipped
    error = db.Column(db.String(500), nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sent_at = db.Column(db.DateTime, nullable=True)


class WhatsappWebhookLog(db.Model):
    """Диагностический журнал входящих webhook-вызовов от Evolution API."""
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    remote_ip = db.Column(db.String(64), nullable=True)
    status_code = db.Column(db.Integer, nullable=False, default=200)
    event = db.Column(db.String(64), nullable=True)
    remote_jid = db.Column(db.String(128), nullable=True)
    digits = db.Column(db.String(32), nullable=True)
    matched_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    note = db.Column(db.String(255), nullable=True)
    raw_body = db.Column(db.Text, nullable=True)


class WorkLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)
    before_photo_url = db.Column(db.String(255), nullable=True)
    after_photo_url = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class Announcement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    text = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class Poll(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    is_anonymous = db.Column(db.Boolean, nullable=False, default=True)
    is_open = db.Column(db.Boolean, nullable=False, default=True)
    result_visibility = db.Column(db.String(20), nullable=False, default="immediate")
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class Vote(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    poll_id = db.Column(db.Integer, db.ForeignKey("poll.id"), nullable=False)
    apartment_id = db.Column(db.Integer, db.ForeignKey("apartment.id"), nullable=False)
    choice = db.Column(db.Enum("yes", "no", name="vote_choice"), nullable=False)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    actor_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    action = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


# Kateqoriyalar (Xərclər): şablon və xərc sətirlərində istifadə olunur.
EXPENSE_CATEGORIES = (
    "əmək haqqı",
    "əlavə hərc",
    "komunal",
    "servis",
)


def _parse_expense_category(raw: Optional[str]) -> Optional[str]:
    v = (raw or "").strip()
    return v if v in EXPENSE_CATEGORIES else None


class ExpenseTemplate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    category = db.Column(db.String(64), nullable=True)
    default_amount = db.Column(db.Numeric(12, 2), nullable=False, default=Decimal("0.00"))
    is_recurring = db.Column(db.Boolean, nullable=False, default=True)
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    period = db.Column(db.String(7), nullable=False)  # YYYY-MM
    name = db.Column(db.String(120), nullable=False)
    category = db.Column(db.String(64), nullable=True)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    is_paid = db.Column(db.Boolean, nullable=False, default=False)
    paid_at = db.Column(db.DateTime, nullable=True)
    template_id = db.Column(db.Integer, db.ForeignKey("expense_template.id"), nullable=True)
    template = db.relationship("ExpenseTemplate", backref="expenses")
    created_by_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class BalanceTopUp(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    comment = db.Column(db.String(255), nullable=True)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


def audit(action: str) -> None:
    user_id = session.get("user_id")
    if user_id:
        db.session.add(AuditLog(actor_user_id=user_id, action=action))
        db.session.commit()


def _apply_credit_to_invoice(invoice: "Invoice") -> float:
    """Apply apartment credit to this invoice, return applied amount."""
    apt = invoice.apartment
    if not apt:
        return 0.0
    credit = float(apt.credit_balance or 0)
    if credit <= 0:
        return 0.0
    remaining = max(0.0, float(invoice.amount or 0) - float(invoice.paid_amount or 0))
    if remaining <= 0:
        return 0.0
    applied = min(credit, remaining)
    invoice.paid_amount = round(float(invoice.paid_amount or 0) + applied, 2)
    apt.credit_balance = round(credit - applied, 2)
    invoice.status = INVOICE_STATUS_PAID if float(invoice.paid_amount or 0) >= float(invoice.amount or 0) else "gozlemede"
    return round(applied, 2)


def _move_invoice_overpay_to_credit(invoice: "Invoice") -> float:
    """If invoice is overpaid, cap paid_amount and move overflow to apartment credit."""
    apt = invoice.apartment
    if not apt:
        return 0.0
    overflow = float(invoice.paid_amount or 0) - float(invoice.amount or 0)
    if overflow <= 0:
        return 0.0
    invoice.paid_amount = float(invoice.amount or 0)
    apt.credit_balance = round(float(apt.credit_balance or 0) + overflow, 2)
    invoice.status = INVOICE_STATUS_PAID
    return round(overflow, 2)


def _apply_payment_delta(invoice: "Invoice", delta: float, *, debt_adjustment: bool = False) -> dict:
    """
    Apply a payment delta to an invoice/apartment.
    Positive delta increases paid; negative delta is a correction (decrease) or,
    when debt_adjustment is True (admin negative Mədaxil), records extra debt
    by lowering paid_amount (may go negative so balance = paid - amount matches).

    Returns dict with keys: moved_to_credit, removed_from_credit.
    """
    apt = invoice.apartment
    if not apt:
        # No apartment linked — just update paid_amount directly.
        invoice.paid_amount = round(float(invoice.paid_amount or 0) + float(delta), 2)
        invoice.status = INVOICE_STATUS_PAID if float(invoice.paid_amount or 0) >= float(invoice.amount or 0) else "gozlemede"
        return {"moved_to_credit": 0.0, "removed_from_credit": 0.0}

    delta = float(delta or 0)
    moved_to_credit = 0.0
    removed_from_credit = 0.0

    if delta >= 0:
        # Payment: add to paid_amount, then push any overflow into apartment credit.
        invoice.paid_amount = round(float(invoice.paid_amount or 0) + delta, 2)
        moved_to_credit = _move_invoice_overpay_to_credit(invoice)
    else:
        if debt_adjustment:
            # Additional debt / manual adjustment: do not pull from apartment credit first.
            invoice.paid_amount = round(float(invoice.paid_amount or 0) + delta, 2)
        else:
            # Correction / reversal: reduce paid_amount.
            # Strategy: first try to recover from apartment credit (so the credit
            # pool shrinks rather than paid_amount going negative), then reduce
            # paid_amount for whatever remains.
            need = -delta  # positive amount we need to subtract
            credit = float(apt.credit_balance or 0)
            take_credit = min(credit, need)
            if take_credit > 0:
                apt.credit_balance = round(credit - take_credit, 2)
                removed_from_credit = round(take_credit, 2)
                need -= take_credit

            if need > 0:
                invoice.paid_amount = max(0.0, round(float(invoice.paid_amount or 0) - need, 2))

        invoice.status = INVOICE_STATUS_PAID if float(invoice.paid_amount or 0) >= float(invoice.amount or 0) else "gozlemede"

    return {
        "moved_to_credit": round(float(moved_to_credit or 0), 2),
        "removed_from_credit": round(float(removed_from_credit or 0), 2),
    }


IMAGE_MAX_BYTES = 5 * 1024 * 1024  # 5 MB на одну картинку
_PIL_FORMAT_TO_EXT = {
    "JPEG": {"jpg", "jpeg"},
    "PNG": {"png"},
    "WEBP": {"webp"},
    "GIF": {"gif"},
}


def save_uploaded_image(file_storage):
    """Сохранить загруженную картинку с валидацией:
    - whitelist расширений (jpg/jpeg/png/webp/gif),
    - лимит 5 MB на файл,
    - реальная проверка содержимого через Pillow (magic bytes + decoding),
    - соответствие реального формата расширению (anti-polyglot).
    При любой ошибке пишет flash и возвращает None.
    """
    if not file_storage or not file_storage.filename:
        return None

    allowed_ext = {"jpg", "jpeg", "png", "webp", "gif"}
    ext = file_storage.filename.rsplit(".", 1)[-1].lower() if "." in file_storage.filename else ""
    if ext not in allowed_ext:
        flash("Şəkil formatı dəstəklənmir (jpg, png, webp, gif).", "warning")
        return None

    stream = file_storage.stream
    try:
        stream.seek(0, 2)
        size = stream.tell()
        stream.seek(0)
    except (OSError, ValueError):
        flash("Şəkil yüklənmədi.", "warning")
        return None

    if size == 0:
        return None
    if size > IMAGE_MAX_BYTES:
        flash("Şəkil ölçüsü 5 MB-dan böyükdür.", "warning")
        return None

    # verify() отмечает объект как непригодный для дальнейших операций,
    # поэтому используем отдельный open() для проверки и потом пере-seek'аем stream для save().
    try:
        with Image.open(stream) as img:
            img.verify()
            detected = (img.format or "").upper()
    except (UnidentifiedImageError, OSError, SyntaxError):
        flash("Fayl yararsız şəkildir.", "warning")
        return None
    finally:
        try:
            stream.seek(0)
        except (OSError, ValueError):
            pass

    if ext not in _PIL_FORMAT_TO_EXT.get(detected, set()):
        flash("Şəkil məzmunu uzantı ilə uyğun deyil.", "warning")
        return None

    uploads_dir = Path(app.root_path) / app.config["UPLOAD_FOLDER"] / "worklog"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    original_name = secure_filename(file_storage.filename)
    unique_name = f"{uuid.uuid4().hex}_{original_name}"
    target_path = uploads_dir / unique_name
    file_storage.save(target_path)
    return f"/static/uploads/worklog/{unique_name}"


def _is_sqlite() -> bool:
    """PRAGMA и нетипичные ALTER TABLE работают только в SQLite.
    На других диалектах схему поддерживает db.create_all() + полноценные миграции."""
    try:
        return db.engine.dialect.name == "sqlite"
    except Exception:
        return False


def ensure_poll_schema():
    # Lightweight SQLite migration for already-created DB.
    if not _is_sqlite():
        return
    with db.engine.connect() as conn:
        columns = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(poll)")}
        if "result_visibility" not in columns:
            conn.exec_driver_sql("ALTER TABLE poll ADD COLUMN result_visibility VARCHAR(20) DEFAULT 'immediate' NOT NULL")


def ensure_payment_schema():
    # Lightweight SQLite migration for already-created DB.
    if not _is_sqlite():
        return
    with db.engine.connect() as conn:
        columns = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(payment)")}
        if "status" not in columns:
            conn.exec_driver_sql("ALTER TABLE payment ADD COLUMN status VARCHAR(20) DEFAULT 'confirmed' NOT NULL")
        if "reviewer_user_id" not in columns:
            conn.exec_driver_sql("ALTER TABLE payment ADD COLUMN reviewer_user_id INTEGER")
        if "comment" not in columns:
            conn.exec_driver_sql("ALTER TABLE payment ADD COLUMN comment VARCHAR(255)")


def ensure_apartment_schema():
    if not _is_sqlite():
        return
    with db.engine.connect() as conn:
        columns = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(apartment)")}
        if "credit_balance" not in columns:
            conn.exec_driver_sql("ALTER TABLE apartment ADD COLUMN credit_balance NUMERIC NOT NULL DEFAULT 0")
        if "rooms" not in columns:
            conn.exec_driver_sql("ALTER TABLE apartment ADD COLUMN rooms INTEGER")


def ensure_apartment_preset_schema():
    inspector = sa_inspect(db.engine)
    if not inspector.has_table("apartment_preset"):
        db.create_all()


def ensure_building_schema():
    """Create building table if missing and add building_id column to apartment."""
    inspector = sa_inspect(db.engine)
    if not inspector.has_table("building"):
        db.create_all()
    # Add building_id to apartment if not present (for existing DBs)
    if not _is_sqlite():
        return
    with db.engine.connect() as conn:
        apartment_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(apartment)")}
        if "building_id" not in apartment_cols:
            conn.exec_driver_sql("ALTER TABLE apartment ADD COLUMN building_id INTEGER REFERENCES building(id)")


def ensure_money_numeric_schema():
    """
    SQLite migration: rebuild tables so monetary fields use NUMERIC instead of FLOAT/REAL.
    Existing values are preserved via CAST(... AS NUMERIC).
    """
    with db.engine.begin() as conn:
        dialect = conn.dialect.name
        if dialect != "sqlite":
            return

        table_defs = {
            "apartment": {
                "money_columns": {"credit_balance"},
                "create_sql": """
                    CREATE TABLE apartment_new (
                        id INTEGER NOT NULL PRIMARY KEY,
                        number VARCHAR(20) NOT NULL UNIQUE,
                        floor INTEGER NOT NULL,
                        area FLOAT NOT NULL,
                        owner_user_id INTEGER NOT NULL,
                        credit_balance NUMERIC(12,2) NOT NULL DEFAULT 0,
                        FOREIGN KEY(owner_user_id) REFERENCES user (id)
                    )
                """,
                "copy_sql": """
                    INSERT INTO apartment_new (id, number, floor, area, owner_user_id, credit_balance)
                    SELECT id, number, floor, area, owner_user_id, CAST(credit_balance AS NUMERIC)
                    FROM apartment
                """,
            },
            "tariff": {
                "money_columns": {"amount"},
                "create_sql": """
                    CREATE TABLE tariff_new (
                        id INTEGER NOT NULL PRIMARY KEY,
                        name VARCHAR(120) NOT NULL,
                        type VARCHAR(20) NOT NULL,
                        amount NUMERIC(12,2) NOT NULL,
                        is_active BOOLEAN NOT NULL
                    )
                """,
                "copy_sql": """
                    INSERT INTO tariff_new (id, name, type, amount, is_active)
                    SELECT id, name, type, CAST(amount AS NUMERIC), is_active
                    FROM tariff
                """,
            },
            "invoice": {
                "money_columns": {"amount", "paid_amount"},
                "create_sql": """
                    CREATE TABLE invoice_new (
                        id INTEGER NOT NULL PRIMARY KEY,
                        apartment_id INTEGER NOT NULL,
                        period VARCHAR(7) NOT NULL,
                        amount NUMERIC(12,2) NOT NULL,
                        paid_amount NUMERIC(12,2) NOT NULL DEFAULT 0,
                        status VARCHAR(20) NOT NULL,
                        created_at DATETIME,
                        FOREIGN KEY(apartment_id) REFERENCES apartment (id)
                    )
                """,
                "copy_sql": """
                    INSERT INTO invoice_new (id, apartment_id, period, amount, paid_amount, status, created_at)
                    SELECT id, apartment_id, period, CAST(amount AS NUMERIC), CAST(paid_amount AS NUMERIC), status, created_at
                    FROM invoice
                """,
            },
            "payment": {
                "money_columns": {"amount"},
                "create_sql": """
                    CREATE TABLE payment_new (
                        id INTEGER NOT NULL PRIMARY KEY,
                        invoice_id INTEGER NOT NULL,
                        amount NUMERIC(12,2) NOT NULL,
                        comment VARCHAR(255),
                        status VARCHAR(20) NOT NULL DEFAULT 'confirmed',
                        reviewer_user_id INTEGER,
                        created_at DATETIME,
                        FOREIGN KEY(invoice_id) REFERENCES invoice (id),
                        FOREIGN KEY(reviewer_user_id) REFERENCES user (id)
                    )
                """,
                "copy_sql": """
                    INSERT INTO payment_new (id, invoice_id, amount, comment, status, reviewer_user_id, created_at)
                    SELECT id, invoice_id, CAST(amount AS NUMERIC), comment, status, reviewer_user_id, created_at
                    FROM payment
                """,
            },
            "expense_template": {
                "money_columns": {"default_amount"},
                "create_sql": """
                    CREATE TABLE expense_template_new (
                        id INTEGER NOT NULL PRIMARY KEY,
                        name VARCHAR(120) NOT NULL,
                        category VARCHAR(64),
                        default_amount NUMERIC(12,2) NOT NULL DEFAULT 0,
                        is_recurring BOOLEAN NOT NULL,
                        is_active BOOLEAN NOT NULL,
                        created_at DATETIME
                    )
                """,
                "copy_sql": """
                    INSERT INTO expense_template_new (id, name, category, default_amount, is_recurring, is_active, created_at)
                    SELECT id, name, category, CAST(default_amount AS NUMERIC), is_recurring, is_active, created_at
                    FROM expense_template
                """,
            },
            "expense": {
                "money_columns": {"amount"},
                "create_sql": """
                    CREATE TABLE expense_new (
                        id INTEGER NOT NULL PRIMARY KEY,
                        period VARCHAR(7) NOT NULL,
                        name VARCHAR(120) NOT NULL,
                        category VARCHAR(64),
                        amount NUMERIC(12,2) NOT NULL,
                        is_paid BOOLEAN NOT NULL DEFAULT 0,
                        paid_at DATETIME,
                        template_id INTEGER,
                        created_by_user_id INTEGER,
                        created_at DATETIME,
                        FOREIGN KEY(template_id) REFERENCES expense_template (id),
                        FOREIGN KEY(created_by_user_id) REFERENCES user (id)
                    )
                """,
                "copy_sql": """
                    INSERT INTO expense_new (id, period, name, category, amount, is_paid, paid_at, template_id, created_by_user_id, created_at)
                    SELECT id, period, name, category, CAST(amount AS NUMERIC), is_paid, paid_at, template_id, created_by_user_id, created_at
                    FROM expense
                """,
            },
            "balance_top_up": {
                "money_columns": {"amount"},
                "create_sql": """
                    CREATE TABLE balance_top_up_new (
                        id INTEGER NOT NULL PRIMARY KEY,
                        amount NUMERIC(12,2) NOT NULL,
                        comment VARCHAR(255),
                        created_by_user_id INTEGER,
                        created_at DATETIME,
                        FOREIGN KEY(created_by_user_id) REFERENCES user (id)
                    )
                """,
                "copy_sql": """
                    INSERT INTO balance_top_up_new (id, amount, comment, created_by_user_id, created_at)
                    SELECT id, CAST(amount AS NUMERIC), comment, created_by_user_id, created_at
                    FROM balance_top_up
                """,
            },
        }

        conn.exec_driver_sql("PRAGMA foreign_keys=OFF")
        try:
            tables = {row[0] for row in conn.exec_driver_sql("SELECT name FROM sqlite_master WHERE type='table'")}
            for table_name, cfg in table_defs.items():
                if table_name not in tables:
                    continue

                pragma_rows = conn.exec_driver_sql(f"PRAGMA table_info({table_name})").fetchall()
                col_type = {row[1]: str(row[2] or "").upper() for row in pragma_rows}
                needs_rebuild = any(
                    ("FLOAT" in col_type.get(col, "") or "REAL" in col_type.get(col, ""))
                    for col in cfg["money_columns"]
                )
                if not needs_rebuild:
                    continue

                conn.exec_driver_sql(cfg["create_sql"])
                conn.exec_driver_sql(cfg["copy_sql"])
                conn.exec_driver_sql(f"DROP TABLE {table_name}")
                conn.exec_driver_sql(f"ALTER TABLE {table_name}_new RENAME TO {table_name}")
        finally:
            conn.exec_driver_sql("PRAGMA foreign_keys=ON")


def ensure_system_schema():
    # Lightweight SQLite migration for already-created DB.
    if not _is_sqlite():
        return
    with db.engine.connect() as conn:
        columns = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(smtp_config)")}
        if "system_name" not in columns:
            conn.exec_driver_sql("ALTER TABLE smtp_config ADD COLUMN system_name VARCHAR(120)")
        if "house_address" not in columns:
            conn.exec_driver_sql("ALTER TABLE smtp_config ADD COLUMN house_address VARCHAR(255)")
        if "commandant_name" not in columns:
            conn.exec_driver_sql("ALTER TABLE smtp_config ADD COLUMN commandant_name VARCHAR(120)")
        if "contact_phone" not in columns:
            conn.exec_driver_sql("ALTER TABLE smtp_config ADD COLUMN contact_phone VARCHAR(50)")
        if "portal_url" not in columns:
            conn.exec_driver_sql("ALTER TABLE smtp_config ADD COLUMN portal_url VARCHAR(255)")
        if "whatsapp_group_url" not in columns:
            conn.exec_driver_sql("ALTER TABLE smtp_config ADD COLUMN whatsapp_group_url VARCHAR(512)")


def ensure_user_role_migration():
    # Rename role values (idempotent). Use begin() so DDL/DML is committed (connect() alone rolls back on exit).
    with db.engine.begin() as conn:
        conn.exec_driver_sql("UPDATE user SET role='komendant' WHERE role='commandant'")
        conn.exec_driver_sql("UPDATE user SET role='admin' WHERE role='superadmin'")


def ensure_invoice_paid_status_spelling():
    """Migrate Invoice.status from legacy ASCII 'odenilib' to ödənilib (idempotent)."""
    try:
        with db.engine.begin() as conn:
            conn.execute(
                sa_text("UPDATE invoice SET status = :paid WHERE status = 'odenilib'"),
                {"paid": INVOICE_STATUS_PAID},
            )
    except Exception:
        pass


def ensure_whatsapp_schema():
    """Создать таблицы whatsapp_config / whatsapp_queue / whatsapp_webhook_log и добавить WA-поля в user."""
    inspector = sa_inspect(db.engine)
    if (
        not inspector.has_table("whatsapp_config")
        or not inspector.has_table("whatsapp_queue")
        or not inspector.has_table("whatsapp_webhook_log")
    ):
        db.create_all()
    if not _is_sqlite():
        return
    with db.engine.connect() as conn:
        user_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(user)")}
        if "whatsapp_connected" not in user_cols:
            conn.exec_driver_sql("ALTER TABLE user ADD COLUMN whatsapp_connected BOOLEAN NOT NULL DEFAULT 0")
        if "whatsapp_connected_at" not in user_cols:
            conn.exec_driver_sql("ALTER TABLE user ADD COLUMN whatsapp_connected_at DATETIME")
        if "whatsapp_jid" not in user_cols:
            conn.exec_driver_sql("ALTER TABLE user ADD COLUMN whatsapp_jid VARCHAR(64)")


_did_role_migration = False
_did_expense_schema = False
_did_money_migration = False
_did_default_admin_seed = False
_did_apartment_schema_migration = False
_did_tariff_scope_schema = False
_did_building_schema = False
_did_whatsapp_schema = False


def ensure_default_admin_seed():
    # Ensure there is always a bootstrap admin in a fresh DB.
    bootstrap = User.query.filter_by(email="admin@emtk.itg.az").first()
    if bootstrap:
        if bootstrap.role != "admin":
            bootstrap.role = "admin"
            db.session.commit()
        return
    if User.query.filter_by(role="admin").first():
        return
    db.session.add(
        User(
            full_name="Admin",
            phone="+000000",
            email="admin@emtk.itg.az",
            password_hash=generate_password_hash("admin"),
            role="admin",
        )
    )
    db.session.commit()


def run_startup_migrations():
    """Create tables and apply idempotent schema patches (local dev, workers, after DB restore)."""
    db.create_all()
    ensure_money_numeric_schema()
    ensure_poll_schema()
    ensure_payment_schema()
    ensure_apartment_schema()
    ensure_apartment_preset_schema()
    ensure_system_schema()
    ensure_invoice_paid_status_spelling()
    ensure_expense_schema()
    ensure_balance_schema()
    ensure_tariff_scope_schema()
    ensure_building_schema()
    ensure_whatsapp_schema()
    ensure_user_role_migration()
    ensure_default_admin_seed()


@app.before_request
def _run_role_migration_once():
    global _did_role_migration
    global _did_expense_schema
    global _did_money_migration
    global _did_default_admin_seed
    global _did_apartment_schema_migration
    global _did_tariff_scope_schema
    global _did_building_schema
    global _did_whatsapp_schema
    # Legacy session after role rename superadmin -> admin
    if session.get("role") == "superadmin":
        session["role"] = "admin"
    # Expense sütunları əvvəl (NUMERIC rebuild üçün category mövcud olsun).
    if not _did_expense_schema:
        try:
            ensure_expense_schema()
        finally:
            _did_expense_schema = True
    if not _did_money_migration:
        try:
            ensure_money_numeric_schema()
        finally:
            _did_money_migration = True
    if not _did_role_migration:
        try:
            ensure_user_role_migration()
        finally:
            _did_role_migration = True
    if not _did_apartment_schema_migration:
        try:
            ensure_apartment_schema()
            ensure_apartment_preset_schema()
        finally:
            _did_apartment_schema_migration = True
    # Runs on every worker (unlike if __name__ == "__main__"); needed for gunicorn / missing tables.
    if not _did_tariff_scope_schema:
        try:
            ensure_tariff_scope_schema()
        finally:
            _did_tariff_scope_schema = True
    if not _did_building_schema:
        try:
            ensure_building_schema()
        finally:
            _did_building_schema = True
    if not _did_whatsapp_schema:
        try:
            ensure_whatsapp_schema()
        finally:
            _did_whatsapp_schema = True
    start_whatsapp_worker()
    if _did_default_admin_seed:
        return
    try:
        ensure_default_admin_seed()
    finally:
        _did_default_admin_seed = True


def ensure_expense_schema():
    """Create expense tables if missing; add columns for legacy DBs (SQLite + PostgreSQL)."""
    inspector = sa_inspect(db.engine)
    if not inspector.has_table("expense_template") or not inspector.has_table("expense"):
        db.create_all()
        return

    expense_cols = {c["name"] for c in inspector.get_columns("expense")}
    et_cols = {c["name"] for c in inspector.get_columns("expense_template")}
    dialect = db.engine.dialect.name

    with db.engine.begin() as conn:
        if "is_paid" not in expense_cols:
            if dialect == "sqlite":
                conn.exec_driver_sql("ALTER TABLE expense ADD COLUMN is_paid BOOLEAN NOT NULL DEFAULT 1")
            else:
                conn.exec_driver_sql("ALTER TABLE expense ADD COLUMN is_paid BOOLEAN NOT NULL DEFAULT TRUE")
        if "paid_at" not in expense_cols:
            paid_type = "TIMESTAMP" if dialect == "postgresql" else "DATETIME"
            conn.exec_driver_sql(f"ALTER TABLE expense ADD COLUMN paid_at {paid_type}")

        # Legacy cleanup (SQLite uses 0/1; PostgreSQL uses TRUE/FALSE).
        if dialect == "sqlite":
            conn.exec_driver_sql("UPDATE expense SET is_paid=0 WHERE is_paid IS NULL")
        else:
            conn.exec_driver_sql("UPDATE expense SET is_paid = FALSE WHERE is_paid IS NULL")

        if "category" not in et_cols:
            conn.exec_driver_sql("ALTER TABLE expense_template ADD COLUMN category VARCHAR(64)")
        if "category" not in expense_cols:
            conn.exec_driver_sql("ALTER TABLE expense ADD COLUMN category VARCHAR(64)")


def ensure_balance_schema():
    inspector = sa_inspect(db.engine)
    if not inspector.has_table("balance_top_up"):
        db.create_all()


def ensure_tariff_scope_schema():
    # Use Inspector so this works on SQLite and PostgreSQL (not only sqlite_master).
    inspector = sa_inspect(db.engine)
    if not inspector.has_table("tariff_apartment"):
        db.create_all()


def compute_invoice_amount(apartment, active_tariffs, scope_map):
    total = 0.0
    for t in active_tariffs:
        scope = scope_map.get(t.id)
        if scope is not None and apartment.id not in scope:
            continue
        tariff_amount = float(t.amount or 0)
        total += tariff_amount * float(apartment.area or 0) if t.type == "per_m2" else tariff_amount
    return round(total, 2)


def active_tariff_names_for_apartment(apartment_id: int, active_tariffs, scope_map) -> list:
    """Names of active tariffs applied to the apartment (same scope rules as compute_invoice_amount)."""
    names = []
    for t in active_tariffs:
        scope = scope_map.get(t.id)
        if scope is not None and apartment_id not in scope:
            continue
        nm = (getattr(t, "name", None) or "").strip()
        names.append(nm if nm else f"#{t.id}")
    return names


def get_smtp_config():
    cfg = SmtpConfig.query.first()
    if not cfg:
        cfg = SmtpConfig()
        db.session.add(cfg)
        db.session.commit()
    return cfg


def sqlite_main_database_path() -> Optional[Path]:
    """Filesystem path to the on-disk SQLite database, or None if not file-based SQLite."""
    if db.engine.dialect.name != "sqlite":
        return None
    database = db.engine.url.database
    if not database or database == ":memory:":
        return None
    p = Path(database)
    return p if p.is_absolute() else (Path.cwd() / p).resolve()


def validate_emtk_sqlite_file(path: Path) -> tuple[bool, str]:
    """Return (ok, error_message) if file is a plausible eMTK SQLite backup."""
    try:
        conn = sqlite3.connect(str(path), timeout=10)
        try:
            rows = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
            ).fetchall()
            tables = {r[0] for r in rows}
        finally:
            conn.close()
    except sqlite3.Error as exc:
        return False, f"SQLite faylı oxunmur: {exc}"
    required = {"user", "apartment", "smtp_config"}
    missing = required - tables
    if missing:
        return False, "Bu fayl eMTK bazasının ehtiyat nüsxəsi kimi tanınmadı (cədvəllər çatışmır)."
    return True, ""


def send_email(subject, body, recipients, html_body=None):
    cfg = get_smtp_config()
    if not cfg.host or not cfg.sender_email or not recipients:
        return False
    try:
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = cfg.sender_email
        msg["To"] = ", ".join(recipients)
        msg.set_content(body)
        if html_body:
            msg.add_alternative(html_body, subtype="html")

        with smtplib.SMTP(cfg.host, cfg.port, timeout=20) as server:
            if cfg.use_tls:
                server.starttls()
            if cfg.username:
                server.login(cfg.username, cfg.password or "")
            server.send_message(msg)
        return True
    except Exception:
        return False


def notify_residents(subject, body):
    recipients = [u.email for u in User.query.filter_by(role="resident").all() if u.email]
    return send_email(subject, body, recipients)


def build_invoice_email(invoice, resident, cfg):
    system_name = (cfg.system_name or "").strip() or "eMTK"
    house_address = (cfg.house_address or "").strip()
    commandant_line = f"{cfg.commandant_name}".strip() if cfg.commandant_name else ""
    phone_line = f"{cfg.contact_phone}".strip() if cfg.contact_phone else ""
    header = f"{system_name}"
    if house_address:
        header += f" | {house_address}"
    subject = f"{system_name} - Hesab-faktura {invoice.period} - {invoice.apartment.number}"
    debt_raw = round(invoice.amount - invoice.paid_amount, 2)
    debt = max(0.0, debt_raw)
    credit = max(0.0, -debt_raw)
    body_lines = [
        header,
        "-" * len(header),
        f"Sakin: {resident.full_name}",
        f"Menzil: {invoice.apartment.number}",
        f"Period: {invoice.period}",
        "",
        f"Hesablanıb: {invoice.amount:.2f} AZN",
        f"ödənilib: {invoice.paid_amount:.2f} AZN",
        f"Borc: {debt:.2f} AZN",
        f"Kredit: {credit:.2f} AZN",
        f"Status: {invoice.status}",
    ]
    if commandant_line or phone_line:
        body_lines += ["", "Elaqe:"]
        if commandant_line:
            body_lines.append(f"Komendant: {commandant_line}")
        if phone_line:
            body_lines.append(f"Telefon: {phone_line}")
    plain_body = "\n".join(body_lines) + "\n"
    html_body = render_template(
        "invoice_email.html",
        invoice=invoice,
        resident=resident,
        cfg=cfg,
        system_name=system_name,
        issue_date=utc_to_local(invoice.created_at),
    )
    return subject, plain_body, html_body


def build_receipt_email(payment, cfg):
    invoice = payment.invoice
    resident = invoice.apartment.owner
    system_name = (cfg.system_name or "").strip() or "eMTK"
    subject = f"{system_name} - Ödəniş qəbzi #{payment.id} - {invoice.apartment.number}"
    balance = float(invoice.paid_amount or 0) - float(invoice.amount or 0)
    plain_body = (
        f"{system_name}\n"
        f"Ödəniş qəbzi #{payment.id}\n"
        f"Sakin: {resident.full_name}\n"
        f"Mənzil: {invoice.apartment.number}\n"
        f"Period: {invoice.period}\n"
        f"Ödəniş: {float(payment.amount):.2f} AZN\n"
        f"Hesablanıb: {float(invoice.amount):.2f} AZN\n"
        f"ödənilib: {float(invoice.paid_amount):.2f} AZN\n"
        f"Balans: {balance:.2f} AZN\n"
        f"Tarix: {(payment.created_at or datetime.now(timezone.utc)).strftime('%d.%m.%Y %H:%M')}\n"
    )
    html_body = render_template("receipt_email.html", payment=payment, cfg=cfg, system_name=system_name)
    return subject, plain_body, html_body


def build_whatsapp_receipt_text(payment, cfg) -> str:
    """Qısa ödəniş qəbzi mətni + portalda qəbz linki (WhatsApp)."""
    invoice = payment.invoice
    resident = invoice.apartment.owner if invoice and invoice.apartment else None
    system_name = (cfg.system_name or "").strip() or "eMTK"
    balance = float(invoice.paid_amount or 0) - float(invoice.amount or 0)
    portal_url = (cfg.portal_url or "").strip().rstrip("/")
    try:
        if portal_url:
            path = url_for("resident_receipt", payment_id=payment.id)
            link = f"{portal_url}{path}"
        else:
            link = url_for("resident_receipt", payment_id=payment.id, _external=True)
    except RuntimeError:
        link = ""

    lines = [
        f"*{system_name}*",
        f"Ödəniş qəbzi #{payment.id}",
        "",
        f"Sakin: {resident.full_name if resident else '-'}",
        f"Mənzil: {invoice.apartment.number}",
        f"Period: {invoice.period}",
        "",
        f"Ödəniş: {float(payment.amount):.2f} AZN",
        f"Hesablanıb: {float(invoice.amount):.2f} AZN",
        f"ödənilib: {float(invoice.paid_amount):.2f} AZN",
        f"Balans: {balance:.2f} AZN",
        f"Tarix: {(payment.created_at or datetime.now(timezone.utc)).strftime('%d.%m.%Y %H:%M')}",
    ]
    if link:
        lines += ["", f"Qəbz: {link}"]
    if portal_url:
        lines += ["", f"Portal: {portal_url}"]
    contact_phone = (cfg.contact_phone or "").strip()
    commandant_name = (cfg.commandant_name or "").strip()
    if contact_phone or commandant_name:
        lines.append("")
        if contact_phone:
            lines.append(f"Əlaqə: {contact_phone}")
        if commandant_name:
            lines.append(commandant_name)
    return "\n".join(lines)


def maybe_enqueue_payment_receipt_whatsapp(payment: "Payment") -> None:
    """Təsdiqlənmiş ödəniş üçün qəbz mətnini WhatsApp növbəsinə əlavə edir (rate limit üçün)."""
    if not payment or (payment.status or "") != "confirmed":
        return
    wa_cfg = get_whatsapp_config()
    if not (wa_cfg.enabled and wa_cfg.api_url and wa_cfg.api_key and wa_cfg.instance):
        return
    invoice = payment.invoice
    if not invoice or not invoice.apartment:
        return
    resident = invoice.apartment.owner
    if not resident or not (resident.phone or "").strip():
        return
    cfg = get_smtp_config()
    text = build_whatsapp_receipt_text(payment, cfg)
    if wa_queue_enqueue(resident.phone, text, invoice_id=invoice.id, user_id=resident.id):
        try:
            wa_queue_drain_once()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# WhatsApp (Evolution API) integration
# ---------------------------------------------------------------------------

def get_whatsapp_config() -> "WhatsappConfig":
    """Singleton-накопитель настроек Evolution API."""
    cfg = WhatsappConfig.query.first()
    if not cfg:
        cfg = WhatsappConfig(webhook_secret=secrets.token_urlsafe(24))
        db.session.add(cfg)
        db.session.commit()
    elif not cfg.webhook_secret:
        cfg.webhook_secret = secrets.token_urlsafe(24)
        db.session.commit()
    return cfg


def _wa_digits(phone: str) -> Optional[str]:
    """Нормализация телефона к виду '994XXXXXXXXX' (без '+')."""
    if not phone:
        return None
    s = re.sub(r"\D+", "", str(phone))
    if len(s) < 8:
        return None
    return s


def wa_send_text(phone: str, text: str) -> tuple[bool, str]:
    """Прямой вызов Evolution API: POST {api_url}/message/sendText/{instance}."""
    cfg = get_whatsapp_config()
    if not cfg.enabled:
        return False, "WhatsApp inteqrasiyası deaktiv edilib."
    if not (cfg.api_url and cfg.api_key and cfg.instance):
        return False, "WhatsApp ayarları tam deyil (URL/API key/instance)."
    number = _wa_digits(phone)
    if not number:
        return False, "Telefon nömrəsi yanlışdır."
    url = f"{cfg.api_url.rstrip('/')}/message/sendText/{cfg.instance}"
    headers = {"apikey": cfg.api_key, "Content-Type": "application/json"}
    payload = {"number": number, "text": text}
    try:
        r = requests.post(url, headers=headers, json=payload, timeout=20)
        if 200 <= r.status_code < 300:
            return True, ""
        return False, f"HTTP {r.status_code}: {r.text[:200]}"
    except requests.RequestException as exc:
        return False, f"Şəbəkə xətası: {exc}"


def build_whatsapp_invoice_text(invoice, resident, smtp_cfg) -> str:
    """Короткий текст + ссылка на печатную версию инвойса (резидентская панель)."""
    system_name = (smtp_cfg.system_name or "").strip() or "eMTK"
    balance = round(float(invoice.amount) - float(invoice.paid_amount), 2)
    portal_url = (smtp_cfg.portal_url or "").strip().rstrip("/")
    # Резидентская ссылка (требует логина) — админский /admin/... сакинам недоступен.
    # Если задан portal_url в настройках — используем его как базу, иначе SERVER_NAME из запроса.
    try:
        if portal_url:
            path = url_for("resident_invoice_print", invoice_id=invoice.id)
            link = f"{portal_url}{path}"
        else:
            link = url_for("resident_invoice_print", invoice_id=invoice.id, _external=True)
    except RuntimeError:
        link = ""

    lines = [
        f"*{system_name}*",
        f"Hesab-faktura {invoice.period}",
        "",
        f"Sakin: {resident.full_name}",
        f"Mənzil: {invoice.apartment.number}",
        "",
        f"Hesablanıb: {float(invoice.amount):.2f} AZN",
        f"ödənilib: {float(invoice.paid_amount):.2f} AZN",
        f"Balans: {balance:.2f} AZN",
    ]
    if link:
        lines += ["", f"Çap versiyası: {link}"]

    if portal_url:
        lines += ["", f"Portal: {portal_url}"]

    contact_phone = (smtp_cfg.contact_phone or "").strip()
    commandant_name = (smtp_cfg.commandant_name or "").strip()
    if contact_phone or commandant_name:
        lines.append("")
        if contact_phone:
            lines.append(f"Əlaqə: {contact_phone}")
        if commandant_name:
            lines.append(commandant_name)

    return "\n".join(lines)


def wa_queue_enqueue(phone: str, text: str, *, invoice_id: Optional[int] = None, user_id: Optional[int] = None) -> Optional["WhatsappQueue"]:
    """Ставит сообщение в очередь на отправку."""
    digits = _wa_digits(phone)
    if not digits:
        return None
    item = WhatsappQueue(
        recipient_phone=digits,
        text=text,
        invoice_id=invoice_id,
        user_id=user_id,
        status="pending",
    )
    db.session.add(item)
    db.session.commit()
    return item


WA_BROADCAST_MAX_LEN = 4096

# Унифицированный шаблон персональных WhatsApp-сообщений:
#   Salam {name},
#
#   {body}
#
#   Hörmətlə, Komendant
WA_GREETING_FALLBACK = "hörmətli sakin"
WA_SIGNATURE = "Hörmətlə, Komendant"


def _wa_greeting_name(full_name: Optional[str]) -> str:
    """Имя для приветствия: full_name или fallback, если пусто."""
    s = (full_name or "").strip()
    return s if s else WA_GREETING_FALLBACK


def wrap_whatsapp_personal_text(full_name: Optional[str], body: str) -> str:
    """
    Оборачивает тело сообщения персональным приветствием и фиксированной
    подписью «Hörmətlə, Komendant». Итоговая длина не превышает
    WA_BROADCAST_MAX_LEN — при необходимости укорачивается тело.
    """
    name = _wa_greeting_name(full_name)
    body = (body or "").strip()
    prefix = f"Salam {name},\n\n"
    suffix = f"\n\n{WA_SIGNATURE}"
    max_body = WA_BROADCAST_MAX_LEN - len(prefix) - len(suffix)
    if max_body < 16:
        max_body = 16
    if len(body) > max_body:
        body = body[: max_body - 1].rstrip() + "…"
    return f"{prefix}{body}{suffix}"


def build_whatsapp_content_broadcast_text(kind: str, title: str, body: str, smtp_cfg) -> str:
    """
    Mesaj gövdəsi (salam və imza olmadan) — elan və ya iş qeydi üçün.
    Salam/imza sonradan `wrap_whatsapp_personal_text` tərəfindən əlavə olunur.
    kind: 'elan' | 'is'.
    """
    label = "Yeni elan" if kind == "elan" else "Yeni iş qeydi"
    t = (title or "").strip()
    b = (body or "").strip()
    parts: list[str] = [label]
    if t:
        parts.append(t)
    if b:
        parts.append(b)
    return "\n\n".join(parts)


def wa_broadcast_enqueue(users: list, text: str) -> tuple[int, int]:
    """
    Sakinlərə eyni mətni növbəyə əlavə edir (telefon üzrə təkrarlar bir dəfə).
    qaytarır: (növbəyə düşən, telefonu olmayan / boş)
    """
    text = (text or "").strip()
    if not text:
        return 0, 0
    if len(text) > WA_BROADCAST_MAX_LEN:
        text = text[: WA_BROADCAST_MAX_LEN - 1] + "…"
    seen_digits: set[str] = set()
    enq, skipped = 0, 0
    for u in users:
        digits = _wa_digits(u.phone or "")
        if not digits:
            skipped += 1
            continue
        if digits in seen_digits:
            continue
        seen_digits.add(digits)
        if wa_queue_enqueue(u.phone, text, user_id=u.id):
            enq += 1
    return enq, skipped


def wa_broadcast_enqueue_personal(users: list, body: str) -> tuple[int, int]:
    """
    Персонализированная рассылка: тело оборачивается шаблоном
    «Salam {name}, … Hörmətlə, Komendant» индивидуально для каждого получателя.
    Возвращает (queued, skipped). Дубли по номеру телефона игнорируются
    (первое вхождение побеждает).
    """
    body = (body or "").strip()
    if not body:
        return 0, 0
    seen_digits: set[str] = set()
    enq, skipped = 0, 0
    for u in users:
        digits = _wa_digits(u.phone or "")
        if not digits:
            skipped += 1
            continue
        if digits in seen_digits:
            continue
        seen_digits.add(digits)
        text = wrap_whatsapp_personal_text(u.full_name, body)
        if wa_queue_enqueue(u.phone, text, user_id=u.id):
            enq += 1
    return enq, skipped


def wa_queue_drain_once() -> int:
    """
    Отправляет не более (bulk_limit - отправленных за последнее окно) сообщений.
    Возвращает количество фактически отправленных (успех+ошибка).
    """
    cfg = get_whatsapp_config()
    if not cfg.enabled:
        return 0

    window_start = datetime.now(timezone.utc) - timedelta(seconds=int(cfg.bulk_window_sec or 300))
    sent_in_window = (
        WhatsappQueue.query
        .filter(WhatsappQueue.status == "sent", WhatsappQueue.sent_at >= window_start)
        .count()
    )
    budget = max(0, int(cfg.bulk_limit or 10) - sent_in_window)
    if budget <= 0:
        return 0

    pending = (
        WhatsappQueue.query
        .filter_by(status="pending")
        .order_by(WhatsappQueue.id.asc())
        .limit(budget)
        .all()
    )
    if not pending:
        return 0

    processed = 0
    for item in pending:
        ok, err = wa_send_text(item.recipient_phone, item.text)
        item.status = "sent" if ok else "failed"
        item.error = None if ok else (err or "")[:500]
        item.sent_at = datetime.now(timezone.utc)
        db.session.commit()
        processed += 1
    return processed


def _wa_worker_loop():
    """Фоновый поток: раз в 30 сек вытягивает очередь в пределах лимита."""
    while True:
        try:
            with app.app_context():
                wa_queue_drain_once()
        except Exception:
            # Не валим поток из-за транзиентных ошибок.
            pass
        _time.sleep(30)


def start_whatsapp_worker():
    if app.config.get("WA_WORKER_STARTED"):
        return
    app.config["WA_WORKER_STARTED"] = True
    t = threading.Thread(target=_wa_worker_loop, name="wa-queue-worker", daemon=True)
    t.start()


@app.context_processor
def inject_system_config():
    cfg = get_smtp_config()
    system_name = (cfg.system_name or "").strip() or "eMTK"
    wa_digits = _wa_digits((cfg.contact_phone or "").strip())
    support_whatsapp_url = f"https://wa.me/{wa_digits}" if wa_digits else None
    wg = (cfg.whatsapp_group_url or "").strip()
    if wg and not wg.lower().startswith(("http://", "https://")):
        wg = "https://" + wg
    whatsapp_group_url = wg[:512] if wg else None
    return {
        "system_name": system_name,
        "support_whatsapp_url": support_whatsapp_url,
        "whatsapp_group_url": whatsapp_group_url,
    }


@app.context_processor
def inject_whatsapp_config():
    try:
        cfg = get_whatsapp_config()
    except Exception:
        return {
            "wa_enabled": False,
            "wa_service_number": None,
            "wa_connect_available": False,
            "wa_resident_hint_no_service": False,
            "current_user_obj": None,
            "wa_bulk_limit": 10,
            "wa_bulk_window_sec": 300,
        }
    user = None
    try:
        user = current_user()
    except Exception:
        user = None
    has_api = bool(cfg.api_url and cfg.api_key and cfg.instance)
    wa_on = bool(cfg.enabled and has_api)
    service_set = bool((cfg.service_number or "").strip())
    return {
        "wa_enabled": wa_on,
        "wa_service_number": cfg.service_number,
        "wa_connect_available": service_set,
        "wa_resident_hint_no_service": bool(wa_on and not service_set),
        "current_user_obj": user,
        "wa_bulk_limit": int(cfg.bulk_limit or 10),
        "wa_bulk_window_sec": int(cfg.bulk_window_sec or 300),
    }


def payment_status_label(status: str) -> str:
    return {
        "pending": "gözləmədə",
        "confirmed": "təsdiqlənib",
        "rejected": "İmtina edilib",
    }.get(status or "", status or "")


def payment_status_badge(status: str) -> str:
    return {
        "pending": "warning",
        "confirmed": "success",
        "rejected": "danger",
    }.get(status or "", "secondary")


def _avatar_initials(full_name: str) -> str:
    parts = [p for p in (full_name or "").strip().split() if p]
    if not parts:
        return "?"
    if len(parts) == 1:
        return parts[0][:2].upper()
    return (parts[0][:1] + parts[1][:1]).upper()


def _avatar_tone(key) -> str:
    """Pick a stable palette tone (a..e) based on a string key."""
    s = str(key or "")
    if not s:
        return "e"
    return "abcde"[sum(ord(c) for c in s) % 5]


@app.context_processor
def inject_helpers():
    return {
        "payment_status_label": payment_status_label,
        "payment_status_badge": payment_status_badge,
        "avatar_initials": _avatar_initials,
        "avatar_tone": _avatar_tone,
    }


def get_selected_apartment(user):
    apartments = Apartment.query.filter_by(owner_user_id=user.id).order_by(Apartment.number).all()
    if not apartments:
        return None, []
    selected_id = session.get("selected_apartment_id")
    selected = next((a for a in apartments if a.id == selected_id), apartments[0])
    session["selected_apartment_id"] = selected.id
    return selected, apartments


def normalize_az_phone(phone_raw: str):
    phone = (phone_raw or "").strip()
    if not phone:
        return None
    return phone if re.fullmatch(r"\+994\d{9}", phone) else None


def _parse_int_field(raw_value, *, min_value: int, max_value: int):
    raw = (raw_value or "").strip()
    if not raw:
        return None
    value = int(raw)
    if value < min_value or value > max_value:
        raise ValueError
    return value


def parse_login_identifier(raw: str):
    """Classify login input: email (contains @) or Azerbaijani phone (+994…)."""
    s = (raw or "").strip()
    if not s:
        return None, None
    if "@" in s:
        return "email", s.lower()
    phone = normalize_az_phone(s)
    if phone:
        return "phone", phone
    return None, None


def current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    return db.session.get(User, user_id)


def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not current_user():
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return wrapper


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user or user.role not in roles:
                flash("Bu emeliyyat ucun icazeniz yoxdur.", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)

        return wrapper

    return decorator


@app.route("/")
def root():
    if current_user():
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
@limiter.limit("10 per minute")
def login():
    if request.method == "POST":
        raw = (request.form.get("login") or request.form.get("email") or "").strip()
        password = request.form["password"]
        kind, value = parse_login_identifier(raw)
        if kind is None:
            if raw and "@" not in raw:
                flash("Telefon formatı +994XXXXXXXXX olmalıdır.", "danger")
            else:
                flash("Telefon və ya email daxil edin.", "danger")
            return render_template("login.html")
        if kind == "email":
            user = User.query.filter_by(email=value).first()
        else:
            user = User.query.filter_by(phone=value).first()
        if user and check_password_hash(user.password_hash, password):
            session.clear()
            session["user_id"] = user.id
            session["role"] = user.role
            session.pop("selected_apartment_id", None)
            return redirect(url_for("dashboard"))
        flash("Telefon/email və ya şifrə yanlışdır.", "danger")
    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
@limiter.limit("5 per minute")
def register():
    if request.method == "POST":
        first_name = (request.form.get("first_name") or "").strip()
        last_name = (request.form.get("last_name") or "").strip()
        full_name = f"{first_name} {last_name}".strip()
        if not first_name:
            flash("Adı daxil edin.", "danger")
            return redirect(url_for("register"))
        if not last_name:
            flash("Soyadı daxil edin.", "danger")
            return redirect(url_for("register"))
        if len(full_name) > 120:
            flash("Ad və soyad birlikdə 120 simvoldan çox ola bilməz.", "danger")
            return redirect(url_for("register"))
        phone = request.form.get("phone", "").strip()
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        password_confirm = request.form.get("password_confirm") or ""

        if not email:
            flash("Email daxil edin.", "danger")
            return redirect(url_for("register"))

        if not phone:
            flash("Telefon daxil edin.", "danger")
            return redirect(url_for("register"))

        phone_norm = normalize_az_phone(phone)
        if not phone_norm:
            flash("Telefon formatı +994XXXXXXXXX olmalıdır.", "danger")
            return redirect(url_for("register"))

        if len(password) < 6:
            flash("Şifrə ən azı 6 simvol olmalıdır.", "danger")
            return redirect(url_for("register"))

        if password != password_confirm:
            flash("Şifrələr üst-üstə düşmür.", "danger")
            return redirect(url_for("register"))

        if User.query.filter_by(email=email).first():
            flash("Bu email ile qeydiyyatdan kecmis istifadeci var.", "warning")
            return redirect(url_for("register"))

        if User.query.filter_by(phone=phone_norm).first():
            flash("Bu telefon nömrəsi ilə qeydiyyatdan keçmiş istifadəçi var.", "warning")
            return redirect(url_for("register"))

        resident = User(
            full_name=full_name,
            phone=phone_norm,
            email=email,
            password_hash=generate_password_hash(password),
            role="resident",
        )
        db.session.add(resident)
        db.session.commit()
        flash("Qeydiyyat tamamlandi. Menzil komendant ve ya administrator terefinden teyin edilir.", "success")
        return redirect(url_for("login"))

    return render_template("register.html")


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    user = current_user()
    if request.method == "POST":
        current_pw = request.form.get("current_password", "")
        new_pw = request.form.get("new_password", "")
        confirm_pw = request.form.get("confirm_password", "")
        if not check_password_hash(user.password_hash, current_pw):
            flash("Mövcud şifrə yanlışdır.", "danger")
            return render_template("change_password.html")
        if len(new_pw) < 6:
            flash("Yeni şifrə ən az 6 simvol olmalıdır.", "danger")
            return render_template("change_password.html")
        if new_pw != confirm_pw:
            flash("Yeni şifrələr uyğun gəlmir.", "danger")
            return render_template("change_password.html")
        user.password_hash = generate_password_hash(new_pw)
        db.session.commit()
        audit("Şifrə dəyişdirildi")
        flash("Şifrə uğurla dəyişdirildi.", "success")
        return redirect(url_for("dashboard"))
    return render_template("change_password.html")


@app.route("/dashboard")
@login_required
def dashboard():
    user = current_user()
    if user.role == "resident":
        apartment, apartments = get_selected_apartment(user)
        invoices = Invoice.query.filter_by(apartment_id=apartment.id).order_by(Invoice.created_at.desc()).all() if apartment else []

        base_debt = sum(float(i.amount) - float(i.paid_amount) for i in invoices)
        credit_balance = float(apartment.credit_balance or 0) if apartment else 0.0
        debt = round(base_debt - credit_balance, 2)

        # Building-wide metrics (read-only for residents).
        # Use the same db.case guard as the admin dashboard to avoid negative
        # per-invoice contributions in the SQL aggregate (consistent formula).
        debt_expr = db.case((Invoice.amount - Invoice.paid_amount > 0, Invoice.amount - Invoice.paid_amount), else_=0.0)
        house_total_debt = db.session.query(db.func.sum(debt_expr)).scalar() or 0
        house_credit_total = db.session.query(db.func.sum(Apartment.credit_balance)).scalar() or 0
        house_total_debt = float(house_total_debt or 0) - float(house_credit_total or 0)
        income_total = (
            db.session.query(db.func.sum(Payment.amount)).filter(Payment.status == "confirmed").scalar() or 0
        )
        topup_total = db.session.query(db.func.sum(BalanceTopUp.amount)).scalar() or 0
        paid_expenses_total = (
            db.session.query(db.func.sum(Expense.amount)).filter(Expense.is_paid == True).scalar() or 0
        )
        unpaid_expenses = db.session.query(db.func.sum(Expense.amount)).filter(Expense.is_paid == False).scalar() or 0
        house_balance = round(float(income_total) + float(topup_total) - float(paid_expenses_total), 2)
        recent_expenses = Expense.query.order_by(Expense.created_at.desc()).limit(100).all()
        receipt_by_invoice = {}
        for i in invoices:
            confirmed = sorted([p for p in i.payments if p.status == "confirmed"], key=lambda x: x.created_at, reverse=True)
            receipt_by_invoice[i.id] = confirmed[0].id if confirmed else None
        payment_history = []
        if apartment:
            payment_history = (
                Payment.query.join(Invoice, Payment.invoice_id == Invoice.id)
                .filter(Invoice.apartment_id == apartment.id)
                .order_by(Payment.created_at.desc())
                .limit(100)
                .all()
            )
        works = WorkLog.query.order_by(WorkLog.created_at.desc()).limit(5).all()
        announcements = Announcement.query.order_by(Announcement.created_at.desc()).limit(8).all()

        return render_template(
            "resident_dashboard.html",
            apartment=apartment,
            apartments=apartments,
            invoices=invoices,
            debt=round(debt, 2),
            credit_balance=round(credit_balance, 2),
            house_total_debt=round(float(house_total_debt or 0), 2),
            house_balance=house_balance,
            unpaid_expenses=unpaid_expenses,
            recent_expenses=recent_expenses,
            receipt_by_invoice=receipt_by_invoice,
            works=works,
            announcements=announcements,
            payment_history=payment_history,
        )

    # Normalize the filter range:
    # - оба значения приводим к aware UTC, чтобы совпадать с Payment.created_at;
    # - to_dt используем как ПРАВУЮ границу (эксклюзивно, начало следующего дня),
    #   иначе теряются записи, созданные в сам день to_date после 00:00.
    from_date_raw = request.args.get("from_date")
    to_date_raw = request.args.get("to_date")

    def _parse_date(value):
        try:
            return datetime.strptime(value, "%Y-%m-%d").date() if value else None
        except ValueError:
            return None

    today = date.today()
    from_date_obj = _parse_date(from_date_raw) or date(today.year, today.month, 1)
    to_date_obj = _parse_date(to_date_raw) or today

    from_dt = datetime.combine(from_date_obj, time.min, tzinfo=timezone.utc)
    to_dt_exclusive = datetime.combine(to_date_obj, time.min, tzinfo=timezone.utc) + timedelta(days=1)
    apartments_count = Apartment.query.count()
    # Total debt should not be reduced by overpayments (credit) inside invoices,
    # but should be reduced by apartment credit balances.
    debt_expr = db.case((Invoice.amount - Invoice.paid_amount > 0, Invoice.amount - Invoice.paid_amount), else_=0.0)
    debt = db.session.query(db.func.sum(debt_expr)).scalar() or 0
    house_credit_total = db.session.query(db.func.sum(Apartment.credit_balance)).scalar() or 0
    debt = float(debt or 0) - float(house_credit_total or 0)
    pending_invoices = Invoice.query.filter(Invoice.status != INVOICE_STATUS_PAID).count()
    recent_logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(10).all()
    dashboard_works = WorkLog.query.order_by(WorkLog.created_at.desc()).limit(6).all()
    dashboard_announcements = Announcement.query.order_by(Announcement.created_at.desc()).limit(6).all()
    period_payments = Payment.query.filter(
        Payment.status == "confirmed", Payment.created_at >= from_dt, Payment.created_at < to_dt_exclusive
    ).order_by(Payment.created_at.asc()).all()
    payment_table = []
    for p in period_payments:
        payment_table.append(
            {
                "apartment": p.invoice.apartment.number,
                "period": p.invoice.period,
                "amount": p.amount,
                "date": p.created_at.strftime("%d.%m.%Y"),
            }
        )
    payments_by_day = {}
    for p in period_payments:
        key = p.created_at.strftime("%Y-%m-%d")
        payments_by_day[key] = payments_by_day.get(key, 0) + float(p.amount)

    period_topups = BalanceTopUp.query.filter(BalanceTopUp.created_at >= from_dt, BalanceTopUp.created_at < to_dt_exclusive).order_by(BalanceTopUp.created_at.asc()).all()
    topups_by_day = {}
    for t in period_topups:
        key = t.created_at.strftime("%Y-%m-%d")
        topups_by_day[key] = topups_by_day.get(key, 0) + float(t.amount)
    for d, v in topups_by_day.items():
        payments_by_day[d] = payments_by_day.get(d, 0) + float(v)

    period_expenses = (
        Expense.query.filter(
            Expense.created_at >= from_dt, Expense.created_at < to_dt_exclusive, Expense.is_paid == True
        )
        .order_by(Expense.created_at.asc())
        .all()
    )
    expenses_by_day = {}
    for e in period_expenses:
        key = e.created_at.strftime("%Y-%m-%d")
        expenses_by_day[key] = expenses_by_day.get(key, 0) + float(e.amount)

    chart_labels = sorted(set(payments_by_day.keys()) | set(expenses_by_day.keys()))
    chart_payments = []
    chart_expenses = []
    chart_balance = []
    running = 0.0
    for d in chart_labels:
        income = float(payments_by_day.get(d, 0.0))
        out = float(expenses_by_day.get(d, 0.0))
        running = round(running + income - out, 2)
        chart_payments.append(round(income, 2))
        chart_expenses.append(round(out, 2))
        chart_balance.append(running)
    # Debt by apartment with credit applied.
    debt_rows = (
        db.session.query(Apartment.id, Apartment.number, db.func.sum(debt_expr), Apartment.credit_balance)
        .join(Invoice, Invoice.apartment_id == Apartment.id)
        .group_by(Apartment.id)
        .all()
    )
    debt_by_apartment = []
    for apt_id, apt_no, inv_debt_sum, credit_bal in debt_rows:
        v = float(inv_debt_sum or 0) - float(credit_bal or 0)
        debt_by_apartment.append((apt_no, round(v, 2)))
    debt_by_apartment.sort(key=lambda x: x[0])

    income_total = db.session.query(db.func.sum(Payment.amount)).filter(Payment.status == "confirmed").scalar() or 0
    topup_total = db.session.query(db.func.sum(BalanceTopUp.amount)).scalar() or 0
    paid_expenses_total = (
        db.session.query(db.func.sum(Expense.amount)).filter(Expense.is_paid == True).scalar() or 0
    )
    unpaid_expenses_total = (
        db.session.query(db.func.sum(Expense.amount)).filter(Expense.is_paid == False).scalar() or 0
    )
    apartment_payments_period = (
        db.session.query(db.func.sum(Payment.amount))
        .filter(Payment.status == "confirmed", Payment.created_at >= from_dt, Payment.created_at < to_dt_exclusive)
        .scalar()
        or 0
    )
    expenses_period = (
        db.session.query(db.func.sum(Expense.amount))
        .filter(Expense.is_paid == True, Expense.created_at >= from_dt, Expense.created_at < to_dt_exclusive)
        .scalar()
        or 0
    )
    house_balance = round(float(income_total) + float(topup_total) - float(paid_expenses_total), 2)

    return render_template(
        "admin_dashboard.html",
        apartments_count=apartments_count,
        total_debt=round(debt, 2),
        house_balance=house_balance,
        apartment_payments_period=round(float(apartment_payments_period or 0), 2),
        expenses_period=round(float(expenses_period or 0), 2),
        paid_expenses_total=round(float(paid_expenses_total or 0), 2),
        unpaid_expenses_total=round(float(unpaid_expenses_total or 0), 2),
        pending_invoices=pending_invoices,
        recent_logs=recent_logs,
        dashboard_works=dashboard_works,
        dashboard_announcements=dashboard_announcements,
        from_date=from_date_obj.strftime("%Y-%m-%d"),
        to_date=to_date_obj.strftime("%Y-%m-%d"),
        payment_table=payment_table,
        chart_labels=chart_labels,
        chart_payments=chart_payments,
        chart_expenses=chart_expenses,
        chart_balance=chart_balance,
        debt_by_apartment=debt_by_apartment,
    )


@app.route("/admin/balance/topup", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def balance_topup():
    try:
        amount = float(request.form.get("amount", "0") or 0)
    except (TypeError, ValueError):
        flash("Məbləğ düzgün deyil.", "danger")
        return redirect(url_for("dashboard"))
    if amount <= 0:
        flash("Məbləğ sıfırdan böyük olmalıdır.", "danger")
        return redirect(url_for("dashboard"))
    comment = (request.form.get("comment", "") or "").strip() or None
    db.session.add(BalanceTopUp(amount=round(amount, 2), comment=comment, created_by_user_id=current_user().id))
    db.session.commit()
    audit(f"Balans artirildi {amount:.2f} AZN" + (f" ({comment})" if comment else ""))
    flash("Balans artirildi.", "success")
    return redirect(url_for("dashboard"))


@app.route("/resident/select-apartment", methods=["POST"])
@login_required
@role_required("resident")
def select_apartment():
    user = current_user()
    if Apartment.query.filter_by(owner_user_id=user.id).count() <= 1:
        return redirect(url_for("dashboard"))
    apartment_id = int(request.form["apartment_id"])
    owned = Apartment.query.filter_by(owner_user_id=user.id, id=apartment_id).first()
    if owned:
        session["selected_apartment_id"] = owned.id
    return redirect(url_for("dashboard"))


def _apartment_row_sort_key(ap: Apartment) -> tuple:
    """Sıra: korpus (ad, id), sonra mənzil nömrəsi (əvvəl rəqəm, sonra mətn)."""
    if ap.building_id is not None and ap.building is not None:
        b_name = (ap.building.name or "").strip().lower()
        b_id = int(ap.building_id)
    else:
        b_name = ""
        b_id = 0
    raw = (ap.number or "").strip()
    try:
        n = int(raw)
        tail = ""
    except ValueError:
        m = re.match(r"^(\d+)(.*)$", raw)
        if m:
            n = int(m.group(1))
            tail = (m.group(2) or "").lower()
        else:
            n = 10**9
            tail = raw.lower()
    return (b_name, b_id, n, tail)


@app.route("/admin/apartments", methods=["GET", "POST"])
@login_required
@role_required("komendant", "admin")
def admin_apartments():
    if request.method == "POST":
        number = request.form["number"].strip()
        if not number or len(number) > 4:
            flash("Nömrə 1-4 simvol olmalıdır.", "danger")
            return redirect(url_for("admin_apartments"))
        try:
            floor = _parse_int_field(request.form.get("floor"), min_value=-999, max_value=999)
        except ValueError:
            flash("Mərtəbə -999 ilə 999 aralığında olmalıdır.", "danger")
            return redirect(url_for("admin_apartments"))
        preset_id_raw = (request.form.get("preset_id", "") or "").strip()
        preset = ApartmentPreset.query.get(int(preset_id_raw)) if preset_id_raw.isdigit() else None
        rooms_raw = (request.form.get("rooms", "") or "").strip()
        try:
            rooms = _parse_int_field(rooms_raw, min_value=1, max_value=999) if rooms_raw else None
        except ValueError:
            flash("Otaq sayı 1-999 aralığında olmalıdır.", "danger")
            return redirect(url_for("admin_apartments"))
        area_raw = (request.form.get("area", "") or "").strip()
        try:
            area = _parse_int_field(area_raw, min_value=1, max_value=9999) if area_raw else None
        except ValueError:
            flash("Sahə 1-9999 aralığında tam ədəd olmalıdır.", "danger")
            return redirect(url_for("admin_apartments"))
        if preset:
            rooms = int(preset.rooms)
            area = int(preset.area)
        if area is None:
            flash("Sahə daxil edilməlidir.", "danger")
            return redirect(url_for("admin_apartments"))
        if floor is None:
            flash("Mərtəbə daxil edilməlidir.", "danger")
            return redirect(url_for("admin_apartments"))
        owner_raw = (request.form.get("owner_user_id") or "").strip()
        if not owner_raw.isdigit():
            flash("Sahib seçilməlidir.", "danger")
            return redirect(url_for("admin_apartments"))
        owner_user_id = int(owner_raw)
        building_id_raw = (request.form.get("building_id", "") or "").strip()
        building_id = int(building_id_raw) if building_id_raw.isdigit() else None
        db.session.add(Apartment(number=number, floor=floor, rooms=rooms, area=area, owner_user_id=owner_user_id, building_id=building_id))
        db.session.commit()
        audit(f"Menzil yaradildi {number}")
        flash("Menzil elave edildi.", "success")
        return redirect(url_for("admin_apartments"))

    building_filter_raw = (request.args.get("building_id", "") or "").strip()
    building_filter_id = int(building_filter_raw) if building_filter_raw.isdigit() else None

    apartments_query = Apartment.query.options(
        joinedload(Apartment.building),
        joinedload(Apartment.owner),
    )
    if building_filter_id:
        apartments_query = apartments_query.filter(Apartment.building_id == building_filter_id)
    apartments = apartments_query.all()

    allowed_apartment_sorts = ("number", "floor", "rooms", "area", "tariff", "owner", "balance")
    sort = (request.args.get("sort") or "").strip().lower()
    sort_dir = (request.args.get("dir") or "asc").strip().lower()
    if sort_dir not in ("asc", "desc"):
        sort_dir = "asc"
    if sort not in allowed_apartment_sorts:
        sort = ""
        sort_dir = "asc"

    residents = User.query.filter_by(role="resident").order_by(db.func.lower(User.full_name), User.id).all()
    apartment_presets = ApartmentPreset.query.order_by(ApartmentPreset.rooms.asc(), ApartmentPreset.area.asc()).all()
    buildings = Building.query.order_by(Building.name.asc()).all()
    debt_expr = db.case((Invoice.amount - Invoice.paid_amount > 0, Invoice.amount - Invoice.paid_amount), else_=0.0)
    debt_rows = (
        db.session.query(Invoice.apartment_id, db.func.sum(debt_expr))
        .group_by(Invoice.apartment_id)
        .all()
    )
    inv_balance_by_apartment_id = {apt_id: float(total or 0) for apt_id, total in debt_rows}
    debt_by_apartment_id = {}
    for a in apartments:
        inv_bal = float(inv_balance_by_apartment_id.get(a.id, 0) or 0)
        credit = float(a.credit_balance or 0)
        debt_by_apartment_id[a.id] = round(inv_bal - credit, 2)
    active_tariffs = Tariff.query.filter_by(is_active=True).order_by(Tariff.id.asc()).all()
    scope_rows_all = TariffApartment.query.all()
    scope_map_all = {}
    for r in scope_rows_all:
        scope_map_all.setdefault(r.tariff_id, set()).add(r.apartment_id)
    tariff_names_by_apartment_id = {
        a.id: active_tariff_names_for_apartment(a.id, active_tariffs, scope_map_all) for a in apartments
    }

    if not sort:
        apartments.sort(key=_apartment_row_sort_key)
    else:
        reverse = sort_dir == "desc"
        if sort == "number":
            apartments.sort(key=_apartment_row_sort_key, reverse=reverse)
        elif sort == "floor":
            apartments.sort(key=lambda a: (a.floor, a.id), reverse=reverse)
        elif sort == "rooms":

            def _rooms_sort_key(ap: Apartment):
                if ap.rooms is None:
                    return (1, 0, ap.id)
                return (0, ap.rooms, ap.id)

            apartments.sort(key=_rooms_sort_key, reverse=reverse)
        elif sort == "area":
            apartments.sort(key=lambda a: (float(a.area or 0), a.id), reverse=reverse)
        elif sort == "tariff":

            def _tariff_sort_key(ap: Apartment):
                names = tariff_names_by_apartment_id.get(ap.id) or []
                joined = ", ".join(names).strip().lower()
                return (0, joined, ap.id) if joined else (1, "", ap.id)

            apartments.sort(key=_tariff_sort_key, reverse=reverse)
        elif sort == "owner":
            apartments.sort(
                key=lambda a: ((a.owner.full_name or "").lower(), a.id),
                reverse=reverse,
            )
        elif sort == "balance":
            apartments.sort(
                key=lambda a: (debt_by_apartment_id.get(a.id, 0), a.id),
                reverse=reverse,
            )

    base_apartments_q = {}
    if building_filter_id:
        base_apartments_q["building_id"] = building_filter_id
    apartments_sort_next = {}
    for c in allowed_apartment_sorts:
        q = dict(base_apartments_q)
        if sort != c:
            q.update(sort=c, dir="asc")
        else:
            q.update(sort=c, dir=("desc" if sort_dir == "asc" else "asc"))
        apartments_sort_next[c] = q
    apartments_sort_preserve_q = {"sort": sort, "dir": sort_dir} if sort else {}
    return render_template(
        "admin_apartments.html",
        apartments=apartments,
        residents=residents,
        apartment_presets=apartment_presets,
        debt_by_apartment_id=debt_by_apartment_id,
        buildings=buildings,
        building_filter_id=building_filter_id,
        tariff_names_by_apartment_id=tariff_names_by_apartment_id,
        apartments_sort=sort,
        apartments_sort_dir=sort_dir,
        apartments_sort_next=apartments_sort_next,
        apartments_sort_preserve_q=apartments_sort_preserve_q,
    )


@app.route("/admin/apartments/owner-whatsapp", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def admin_apartment_owner_whatsapp():
    wa_cfg = get_whatsapp_config()
    if not (wa_cfg.enabled and wa_cfg.api_url and wa_cfg.api_key and wa_cfg.instance):
        flash("WhatsApp inteqrasiyası deaktiv və ya ayarları tam deyil.", "warning")
        return _redirect_admin_apartments()
    try:
        uid = int(request.form.get("owner_user_id") or "0")
    except (TypeError, ValueError):
        flash("Yanlış istifadəçi.", "danger")
        return _redirect_admin_apartments()
    owner = User.query.get_or_404(uid)
    if not owner.phone:
        flash("Telefon nömrəsi yoxdur.", "warning")
        return _redirect_admin_apartments()
    if not db.session.scalar(select(exists().where(Apartment.owner_user_id == owner.id))):
        flash("Bu istifadəçi mənzil sahibi kimi təyin edilməyib.", "danger")
        return _redirect_admin_apartments()
    message = (request.form.get("message") or "").strip()
    if not message:
        flash("Mesaj mətni boş ola bilməz.", "danger")
        return _redirect_admin_apartments()
    if len(message) > WA_BROADCAST_MAX_LEN:
        flash(f"Mesaj çox uzundur (maks. {WA_BROADCAST_MAX_LEN} simvol).", "danger")
        return _redirect_admin_apartments()
    text = wrap_whatsapp_personal_text(owner.full_name, message)
    ok, err = wa_send_text(owner.phone, text)
    if ok:
        audit(f"WhatsApp fərdi mesaj göndərildi: sakin #{owner.id} ({owner.email})")
        flash("Mesaj WhatsApp ilə göndərildi.", "success")
    else:
        flash(f"WhatsApp göndərilmədi: {err}", "danger")
    return _redirect_admin_apartments()


def _redirect_admin_apartments():
    bid_raw = (request.form.get("building_id") or "").strip()
    try:
        bid = int(bid_raw) if bid_raw else None
    except ValueError:
        bid = None
    if bid:
        return redirect(url_for("admin_apartments", building_id=bid))
    return redirect(url_for("admin_apartments"))


@app.route("/admin/apartments/delete/<int:apartment_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def delete_apartment(apartment_id):
    apartment = Apartment.query.get_or_404(apartment_id)
    # SQLAlchemy 2.x: avoid legacy Query.count() (can raise in some setups); use EXISTS + scalar().
    has_invoices = bool(
        db.session.scalar(select(exists().where(Invoice.apartment_id == apartment.id)))
    )
    if has_invoices:
        flash("Menzili silmek olmur: bagli hesab var.", "warning")
        return redirect(url_for("admin_apartments"))

    apartment_number = apartment.number
    apt_pk = apartment.id
    # Core DELETEs: avoid ORM session.delete(apartment), which can touch relationships / flush order.
    try:
        # Polls are deprecated; cleanup orphan votes before apartment delete.
        db.session.execute(sa_delete(Vote).where(Vote.apartment_id == apt_pk))
        db.session.execute(sa_delete(TariffApartment).where(TariffApartment.apartment_id == apt_pk))
        db.session.execute(sa_delete(Apartment).where(Apartment.id == apt_pk))
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        app.logger.exception("delete_apartment integrity error apartment_id=%s", apartment_id)
        flash("Menzil silinmədi: bağlı qeydlər var (məsələn, hesab).", "danger")
        return redirect(url_for("admin_apartments"))
    except SQLAlchemyError:
        db.session.rollback()
        app.logger.exception("delete_apartment database error apartment_id=%s", apartment_id)
        flash("Menzil silinmədi: verilənlər bazası xətası.", "danger")
        return redirect(url_for("admin_apartments"))
    except Exception:
        db.session.rollback()
        app.logger.exception("delete_apartment unexpected error apartment_id=%s", apartment_id)
        flash("Menzil silinmədi: gözlənilməz xəta.", "danger")
        return redirect(url_for("admin_apartments"))

    try:
        audit(f"Menzil silindi {apartment_number}")
    except Exception:
        # Delete already committed; avoid 500 if audit log insert fails.
        app.logger.exception("audit after delete_apartment failed apartment_id=%s", apartment_id)
    flash("Menzil silindi.", "success")
    return redirect(url_for("admin_apartments"))


@app.route("/admin/apartments/update/<int:apartment_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def update_apartment(apartment_id):
    apartment = Apartment.query.get_or_404(apartment_id)
    number = request.form["number"].strip()
    if not number or len(number) > 4:
        flash("Nömrə 1-4 simvol olmalıdır.", "danger")
        return redirect(url_for("admin_apartments"))
    try:
        floor = _parse_int_field(request.form.get("floor"), min_value=-999, max_value=999)
    except ValueError:
        flash("Mərtəbə -999 ilə 999 aralığında olmalıdır.", "danger")
        return redirect(url_for("admin_apartments"))
    rooms_raw = (request.form.get("rooms", "") or "").strip()
    try:
        rooms = _parse_int_field(rooms_raw, min_value=1, max_value=999) if rooms_raw else None
    except ValueError:
        flash("Otaq sayı 1-999 aralığında olmalıdır.", "danger")
        return redirect(url_for("admin_apartments"))
    try:
        area = _parse_int_field(request.form.get("area"), min_value=1, max_value=9999)
    except ValueError:
        flash("Sahə 1-9999 aralığında tam ədəd olmalıdır.", "danger")
        return redirect(url_for("admin_apartments"))
    if area is None:
        flash("Sahə daxil edilməlidir.", "danger")
        return redirect(url_for("admin_apartments"))
    owner_user_id = int(request.form["owner_user_id"])

    duplicate = Apartment.query.filter(Apartment.number == number, Apartment.id != apartment.id).first()
    if duplicate:
        flash("Bu nomre ile menzil artiq movcuddur.", "warning")
        return redirect(url_for("admin_apartments"))

    building_id_raw = (request.form.get("building_id", "") or "").strip()
    building_id = int(building_id_raw) if building_id_raw.isdigit() else None

    apartment.number = number
    apartment.floor = floor
    apartment.rooms = rooms
    apartment.area = area
    apartment.owner_user_id = owner_user_id
    apartment.building_id = building_id
    db.session.commit()
    audit(f"Menzil yenilendi {number}")
    flash("Menzil yenilendi.", "success")
    return redirect(url_for("admin_apartments"))


@app.route("/admin/tariffs", methods=["GET", "POST"])
@login_required
@role_required("komendant", "admin")
def admin_tariffs():
    if request.method == "POST":
        is_active = (request.form.get("is_active") or "1").strip() == "1"
        tariff = Tariff(
            name=request.form["name"].strip(),
            type=request.form["type"],
            amount=float(request.form["amount"]),
            is_active=is_active,
        )
        db.session.add(tariff)
        db.session.commit()

        apply_all = request.form.get("apply_all") == "on"
        if not apply_all:
            apartment_ids = request.form.getlist("apartment_ids")
            for apt_id in apartment_ids:
                db.session.add(TariffApartment(tariff_id=tariff.id, apartment_id=int(apt_id)))
            db.session.commit()

        audit("Tarif elave edildi")
        flash("Tarif elave edildi.", "success")
        return redirect(url_for("admin_tariffs"))

    tariffs = Tariff.query.order_by(Tariff.id.desc()).all()
    apartments = Apartment.query.order_by(Apartment.number).all()
    scope_rows = TariffApartment.query.all()
    scope_map = {}
    for r in scope_rows:
        scope_map.setdefault(r.tariff_id, set()).add(r.apartment_id)
    return render_template("admin_tariffs.html", tariffs=tariffs, apartments=apartments, scope_map=scope_map)


@app.route("/admin/tariffs/delete/<int:tariff_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def delete_tariff(tariff_id):
    tariff = Tariff.query.get_or_404(tariff_id)
    tariff_name = tariff.name
    TariffApartment.query.filter_by(tariff_id=tariff.id).delete(synchronize_session=False)
    db.session.delete(tariff)
    db.session.commit()
    audit(f"Tarif silindi {tariff_name}")
    flash("Tarif silindi.", "success")
    return redirect(url_for("admin_tariffs"))


@app.route("/admin/tariffs/update/<int:tariff_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def update_tariff(tariff_id):
    tariff = Tariff.query.get_or_404(tariff_id)

    name = (request.form.get("name", "") or "").strip()
    tariff_type = (request.form.get("type", "") or "").strip()
    amount_raw = (request.form.get("amount", "") or "").strip()
    is_active = request.form.get("is_active") == "on"

    if not name:
        flash("Tarif adı boş ola bilməz.", "danger")
        return redirect(url_for("admin_tariffs"))
    if tariff_type not in ("per_m2", "fixed"):
        flash("Tarif tipi düzgün seçilməyib.", "danger")
        return redirect(url_for("admin_tariffs"))
    try:
        amount = float(amount_raw)
    except ValueError:
        flash("Məbləğ düzgün daxil edilməyib.", "danger")
        return redirect(url_for("admin_tariffs"))
    if amount <= 0:
        flash("Məbləğ sıfırdan böyük olmalıdır.", "danger")
        return redirect(url_for("admin_tariffs"))

    tariff.name = name
    tariff.type = tariff_type
    tariff.amount = amount
    tariff.is_active = is_active

    apply_all = request.form.get("apply_all") == "on"
    TariffApartment.query.filter_by(tariff_id=tariff.id).delete(synchronize_session=False)
    if not apply_all:
        apartment_ids = request.form.getlist("apartment_ids")
        for apt_id in apartment_ids:
            if apt_id.isdigit():
                db.session.add(TariffApartment(tariff_id=tariff.id, apartment_id=int(apt_id)))

    db.session.commit()
    audit(f"Tarif yenilendi {tariff.name}")
    flash("Tarif yeniləndi.", "success")
    return redirect(url_for("admin_tariffs"))


@app.route("/admin/invoices/generate", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def generate_invoices():
    period_mode = (request.form.get("period_mode") or "current").strip().lower()
    today = date.today()
    current_period = today.strftime("%Y-%m")
    next_period = f"{(today.year + (1 if today.month == 12 else 0)):04d}-{(1 if today.month == 12 else today.month + 1):02d}"
    periods_to_generate = [current_period] if period_mode != "next" else [current_period, next_period]

    active_tariffs = Tariff.query.filter_by(is_active=True).all()
    apartments = Apartment.query.all()
    scope_rows = TariffApartment.query.all()
    scope_map = {}
    for r in scope_rows:
        scope_map.setdefault(r.tariff_id, set()).add(r.apartment_id)

    created_by_period = {}
    credit_applied_by_period = {}

    for period in periods_to_generate:
        created = 0
        for apartment in apartments:
            exists = Invoice.query.filter_by(apartment_id=apartment.id, period=period).first()
            if exists:
                continue
            total = compute_invoice_amount(apartment, active_tariffs, scope_map)
            db.session.add(Invoice(apartment_id=apartment.id, period=period, amount=total, status="gozlemede"))
            created += 1
        db.session.commit()
        created_by_period[period] = created

        # Apply apartment credit to all unpaid invoices for the period.
        unpaid_invoices = (
            Invoice.query.join(Apartment, Invoice.apartment_id == Apartment.id)
            .filter(Invoice.period == period, Invoice.status != INVOICE_STATUS_PAID)
            .order_by(Apartment.number.asc(), Invoice.id.asc())
            .all()
        )
        applied_total = 0.0
        for inv in unpaid_invoices:
            applied_total += _apply_credit_to_invoice(inv)
        if applied_total > 0:
            db.session.commit()
            audit(f"Kredit avtomatik tetbiq olundu: {applied_total:.2f} AZN period {period}")
        credit_applied_by_period[period] = round(applied_total, 2)

        templates = ExpenseTemplate.query.filter_by(is_active=True, is_recurring=True).all()
        created_exp = 0
        for t in templates:
            exists = Expense.query.filter_by(period=period, template_id=t.id).first()
            if exists:
                continue
            amount = float(t.default_amount or 0)
            if amount <= 0:
                continue
            db.session.add(
                Expense(
                    period=period,
                    name=t.name,
                    category=t.category,
                    amount=round(amount, 2),
                    is_paid=False,
                    paid_at=None,
                    template_id=t.id,
                    created_by_user_id=current_user().id,
                )
            )
            created_exp += 1
        if created_exp:
            db.session.commit()
            audit(f"Aylıq xərclər yaradıldı: {created_exp} period {period}")
    if len(periods_to_generate) == 2:
        audit(
            f"Hesablar yaradildi ardicil periodlar: {periods_to_generate[0]}={created_by_period.get(periods_to_generate[0], 0)}, "
            f"{periods_to_generate[1]}={created_by_period.get(periods_to_generate[1], 0)}"
        )
        flash(
            "Hesablar yaradildi. "
            f"{periods_to_generate[0]}: {created_by_period.get(periods_to_generate[0], 0)} hesab, "
            f"{periods_to_generate[1]}: {created_by_period.get(periods_to_generate[1], 0)} hesab.",
            "success",
        )
    else:
        period = periods_to_generate[0]
        audit(f"Hesablar yaradildi: {created_by_period.get(period, 0)} period {period}")
        flash(f"Hesablar yaradildi: {created_by_period.get(period, 0)} eded.", "success")

    if any((credit_applied_by_period.get(p, 0) > 0 for p in periods_to_generate)):
        db.session.commit()
    return redirect(url_for("admin_invoices"))


@app.route("/admin/invoices/recalculate", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def recalculate_invoices():
    period = (request.form.get("period") or "").strip()
    if not period or len(period) != 7:
        flash("Period duzgun deyil (YYYY-MM).", "danger")
        return redirect(url_for("admin_tariffs"))

    active_tariffs = Tariff.query.filter_by(is_active=True).all()
    apartments = Apartment.query.all()
    scope_rows = TariffApartment.query.all()
    scope_map = {}
    for r in scope_rows:
        scope_map.setdefault(r.tariff_id, set()).add(r.apartment_id)

    updated = 0
    created = 0
    for apartment in apartments:
        total = compute_invoice_amount(apartment, active_tariffs, scope_map)
        inv = Invoice.query.filter_by(apartment_id=apartment.id, period=period).first()
        if not inv:
            db.session.add(Invoice(apartment_id=apartment.id, period=period, amount=total, status="gozlemede"))
            created += 1
            continue
        inv.amount = total
        inv.status = INVOICE_STATUS_PAID if float(inv.paid_amount or 0) >= float(inv.amount or 0) else "gozlemede"
        overflow = _move_invoice_overpay_to_credit(inv)
        if overflow > 0:
            audit(f"Kredit yarandi (yeniden hesabla) invoice#{inv.id}: +{overflow:.2f} AZN")
        updated += 1

    db.session.commit()

    # FIX (warning #3): After recalculating amounts, apply any accumulated
    # apartment credit to unpaid invoices in this period. Previously the
    # credit was left untouched even if the new amount could be fully or
    # partially covered by it.
    unpaid_invoices = (
        Invoice.query.join(Apartment, Invoice.apartment_id == Apartment.id)
        .filter(Invoice.period == period, Invoice.status != INVOICE_STATUS_PAID)
        .order_by(Apartment.number.asc(), Invoice.id.asc())
        .all()
    )
    applied_total = sum(_apply_credit_to_invoice(inv) for inv in unpaid_invoices)
    if applied_total > 0:
        db.session.commit()
        audit(f"Kredit avtomatik tetbiq olundu (yeniden hesabla): {applied_total:.2f} AZN period {period}")

    audit(f"Hesablar yeniden hesablandi period {period}: updated {updated}, created {created}")
    flash(f"Yeniden hesablandi: {updated}, yaradildi: {created}.", "success")
    return redirect(url_for("admin_invoices"))


@app.route("/admin/expenses", methods=["GET", "POST"])
@login_required
@role_required("komendant", "admin")
def admin_expenses():
    if request.method == "POST":
        form_type = request.form.get("form_type", "")
        if form_type == "add_template":
            name = request.form.get("name", "").strip()
            category = _parse_expense_category(request.form.get("category"))
            default_amount = float(request.form.get("default_amount", "0") or 0)
            # Aylıq şablon UI-də yalnız “Aylıq” seçimi ilə əlavə olunur — həmişə təkrarlanan.
            is_recurring = True
            if not name:
                flash("Ad bos ola bilmez.", "danger")
                return redirect(url_for("admin_expenses"))
            if not category:
                flash("Kateqoriya seçin.", "danger")
                return redirect(url_for("admin_expenses"))
            db.session.add(
                ExpenseTemplate(
                    name=name,
                    category=category,
                    default_amount=round(default_amount, 2),
                    is_recurring=is_recurring,
                    is_active=True,
                )
            )
            db.session.commit()
            audit(f"Xərc şablonu yaradıldı: {name}")
            flash("Xərc şablonu əlavə edildi.", "success")
            return redirect(url_for("admin_expenses"))
        if form_type == "toggle_template":
            template_id = int(request.form["template_id"])
            t = ExpenseTemplate.query.get_or_404(template_id)
            t.is_active = not t.is_active
            db.session.commit()
            audit(f"Xərc şablonu statusu dəyişdi #{template_id}")
            return redirect(url_for("admin_expenses"))
        if form_type == "update_template":
            template_id = int(request.form["template_id"])
            t = ExpenseTemplate.query.get_or_404(template_id)
            name = (request.form.get("name", "") or "").strip()
            category = _parse_expense_category(request.form.get("category"))
            default_amount = float(request.form.get("default_amount", "0") or 0)
            is_recurring = True
            if not name:
                flash("Ad bos ola bilmez.", "danger")
                return redirect(url_for("admin_expenses"))
            if not category:
                flash("Kateqoriya seçin.", "danger")
                return redirect(url_for("admin_expenses"))
            if default_amount < 0:
                flash("Məbləğ mənfi ola bilməz.", "danger")
                return redirect(url_for("admin_expenses"))
            t.name = name
            t.category = category
            t.default_amount = round(default_amount, 2)
            t.is_recurring = is_recurring
            db.session.commit()
            audit(f"Xərc şablonu yeniləndi #{template_id}: {name}")
            flash("Xərc şablonu yeniləndi.", "success")
            return redirect(url_for("admin_expenses"))
        if form_type == "delete_template":
            template_id = int(request.form["template_id"])
            t = ExpenseTemplate.query.get_or_404(template_id)
            Expense.query.filter_by(template_id=t.id).update({"template_id": None}, synchronize_session=False)
            db.session.delete(t)
            db.session.commit()
            audit(f"Xərc şablonu silindi #{template_id}")
            flash("Xərc şablonu silindi.", "success")
            return redirect(url_for("admin_expenses"))
        if form_type == "add_expense":
            period = (request.form.get("period", "") or "").strip()
            name = (request.form.get("name", "") or "").strip()
            category = _parse_expense_category(request.form.get("category"))
            amount = float(request.form.get("amount", "0") or 0)
            if not period or len(period) != 7:
                flash("Period duzgun deyil (YYYY-MM).", "danger")
                return redirect(url_for("admin_expenses"))
            if not name:
                flash("Ad bos ola bilmez.", "danger")
                return redirect(url_for("admin_expenses"))
            if not category:
                flash("Kateqoriya seçin.", "danger")
                return redirect(url_for("admin_expenses"))
            if amount <= 0:
                flash("Məbləğ sıfırdan böyük olmalıdır.", "danger")
                return redirect(url_for("admin_expenses"))
            db.session.add(
                Expense(
                    period=period,
                    name=name,
                    category=category,
                    amount=round(amount, 2),
                    is_paid=False,
                    paid_at=None,
                    created_by_user_id=current_user().id,
                )
            )
            db.session.commit()
            audit(f"Xərc daxil edildi: {name} {amount:.2f} AZN period {period}")
            flash("Xərc daxil edildi.", "success")
            return redirect(url_for("admin_expenses"))

    period = request.args.get("period") or date.today().strftime("%Y-%m")
    sort = (request.args.get("sort") or "").strip().lower()
    sort_dir = (request.args.get("dir") or "asc").strip().lower()
    if sort_dir not in ("asc", "desc"):
        sort_dir = "asc"
    if sort not in ALLOWED_EXPENSE_SORT_KEYS:
        sort = ""
        sort_dir = "asc"

    data = _get_admin_expenses_view_data(period)
    expense_rows = _build_admin_expense_table_rows(
        data["templates"],
        data["one_off_expenses"],
        data["template_expense_by_template_id"],
    )
    if sort:
        _sort_admin_expense_table_rows(expense_rows, sort, sort_dir)

    base_q = {"period": period}
    expenses_sort_next = {}
    for c in ("tip", "tarix", "category", "ad", "mebleg", "sablon", "odenis"):
        q = dict(base_q)
        if sort != c:
            q.update(sort=c, dir="asc")
        else:
            q.update(sort=c, dir=("desc" if sort_dir == "asc" else "asc"))
        expenses_sort_next[c] = q

    data["expense_rows"] = expense_rows
    data["expenses_sort"] = sort
    data["expenses_sort_dir"] = sort_dir
    data["expenses_sort_next"] = expenses_sort_next
    return render_template("admin_expenses.html", **data)


@app.route("/admin/expenses/update/<int:expense_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def update_expense(expense_id):
    e = Expense.query.get_or_404(expense_id)
    period = (request.form.get("period", "") or "").strip()
    name = (request.form.get("name", "") or "").strip()
    category = _parse_expense_category(request.form.get("category"))
    amount = float(request.form.get("amount", "0") or 0)
    if not period or len(period) != 7:
        flash("Period duzgun deyil (YYYY-MM).", "danger")
        return redirect(url_for("admin_expenses", period=e.period))
    if not name:
        flash("Ad bos ola bilmez.", "danger")
        return redirect(url_for("admin_expenses", period=e.period))
    if not category:
        flash("Kateqoriya seçin.", "danger")
        return redirect(url_for("admin_expenses", period=e.period))
    if amount <= 0:
        flash("Məbləğ sıfırdan böyük olmalıdır.", "danger")
        return redirect(url_for("admin_expenses", period=e.period))

    old = f"{e.period} {e.name} {float(e.amount):.2f}"
    period_changed = period != e.period
    e.period = period
    e.name = name
    e.category = category
    e.amount = round(amount, 2)
    # Если период сменили, синхронизируем created_at так, чтобы расход
    # попадал в тот же месяц на странице «Tarixçə» (фильтр по created_at),
    # что и на странице «Xərclər» (фильтр по period).
    if period_changed:
        try:
            target_year, target_month = (int(x) for x in period.split("-"))
        except ValueError:
            target_year = target_month = None
        if target_year and target_month:
            base = e.created_at or datetime.now(timezone.utc)
            if base.tzinfo is None:
                base = base.replace(tzinfo=timezone.utc)
            # День ограничиваем последним днём целевого месяца.
            last_day = calendar.monthrange(target_year, target_month)[1]
            e.created_at = base.replace(
                year=target_year, month=target_month, day=min(base.day, last_day)
            )
            # paid_at тоже согласуем, если расход уже помечен оплаченным.
            if e.is_paid and e.paid_at is not None:
                paid_base = e.paid_at
                if paid_base.tzinfo is None:
                    paid_base = paid_base.replace(tzinfo=timezone.utc)
                e.paid_at = paid_base.replace(
                    year=target_year, month=target_month, day=min(paid_base.day, last_day)
                )
    db.session.commit()
    audit(f"Xərc yeniləndi #{e.id}: {old} -> {e.period} {e.name} {float(e.amount):.2f}")
    flash("Xərc yeniləndi.", "success")
    return redirect(url_for("admin_expenses", period=period))


@app.route("/admin/expenses/delete/<int:expense_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def delete_expense(expense_id):
    e = Expense.query.get_or_404(expense_id)
    period = e.period
    desc = f"{e.name} {float(e.amount):.2f} period {e.period}"
    db.session.delete(e)
    db.session.commit()
    audit(f"Xərc silindi #{expense_id}: {desc}")
    flash("Xərc silindi.", "success")
    return redirect(url_for("admin_expenses", period=period))


@app.route("/admin/expenses/toggle-paid/<int:expense_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def toggle_expense_paid(expense_id):
    e = Expense.query.get_or_404(expense_id)
    e.is_paid = not bool(e.is_paid)
    e.paid_at = datetime.now(timezone.utc) if e.is_paid else None
    db.session.commit()
    audit(f"Xərc {'ödəndi' if e.is_paid else 'ödənilmədi'} #{e.id}: {e.name} {float(e.amount):.2f} period {e.period}")
    flash("Xərc statusu yeniləndi.", "success")
    return redirect(url_for("admin_expenses", period=e.period))


@app.route("/admin/invoices")
@login_required
@role_required("komendant", "admin")
def admin_invoices():
    selected_period = (request.args.get("period", "") or "").strip()
    building_id_raw = (request.args.get("building_id") or "").strip()
    building_id = int(building_id_raw) if building_id_raw.isdigit() else None
    return render_template("admin_invoices.html", **_get_admin_invoices_view_data(selected_period, building_id))


@app.route("/admin/payments/add/<int:invoice_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def add_payment(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    try:
        amount = float(request.form["amount"])
    except (TypeError, ValueError):
        flash("Məbləğ düzgün deyil.", "danger")
        period = (request.form.get("period", "") or "").strip()
        return redirect(url_for("admin_invoices", period=period) if period else url_for("admin_invoices"))
    comment = (request.form.get("comment", "") or "").strip() or None

    if amount == 0:
        flash("Məbləğ sıfır ola bilməz.", "danger")
        period = (request.form.get("period", "") or "").strip()
        return redirect(url_for("admin_invoices", period=period) if period else url_for("admin_invoices"))

    apply_amount = amount
    now = datetime.now(timezone.utc)

    # Оборачиваем мутацию баланса + создание Payment в общий try/except,
    # чтобы исключение не оставило in-memory объекты изменёнными без коммита.
    try:
        result = _apply_payment_delta(
            invoice, apply_amount, debt_adjustment=(apply_amount < 0)
        )
        payment = Payment(
            invoice_id=invoice.id,
            amount=apply_amount,
            comment=comment,
            status="confirmed",
            reviewer_user_id=current_user().id,
            created_at=now,
        )
        db.session.add(payment)
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash("Xəta baş verdi. Ödəniş daxil edilmədi.", "danger")
        period = (request.form.get("period", "") or "").strip()
        return redirect(url_for("admin_invoices", period=period) if period else url_for("admin_invoices"))

    moved = float(result.get("moved_to_credit") or 0)
    removed = float(result.get("removed_from_credit") or 0)
    if moved > 0:
        audit(f"Odenis daxil edildi invoice#{invoice.id} {apply_amount:.2f} AZN (kredit +{moved:.2f})")
    elif removed > 0:
        audit(f"Odenis daxil edildi invoice#{invoice.id} {apply_amount:.2f} AZN (kredit -{removed:.2f})")
    else:
        audit(f"Odenis daxil edildi invoice#{invoice.id} {apply_amount:.2f} AZN")
    resident = invoice.apartment.owner
    if resident and resident.email:
        cfg = get_smtp_config()
        subject, body, html_body = build_receipt_email(payment, cfg)
        send_email(subject, body, [resident.email], html_body=html_body)
    maybe_enqueue_payment_receipt_whatsapp(payment)
    flash("Odenis daxil edildi.", "success")
    period = (request.form.get("period", "") or "").strip()
    return redirect(url_for("admin_invoices", period=period) if period else url_for("admin_invoices"))


@app.route("/admin/history")
@login_required
@role_required("komendant", "admin")
def admin_history():
    selected_month = (request.args.get("month") or "").strip()
    building_id_raw = (request.args.get("building_id") or "").strip()
    building_id = int(building_id_raw) if building_id_raw.isdigit() else None
    return render_template("admin_history.html", **_get_admin_history_view_data(selected_month, building_id))


def _iter_months_inclusive(from_period: str, to_period: str):
    def parse(p):
        y, m = p.split("-")
        return int(y), int(m)

    fy, fm = parse(from_period)
    ty, tm = parse(to_period)
    y, m = fy, fm
    while (y, m) <= (ty, tm):
        yield f"{y:04d}-{m:02d}"
        m += 1
        if m == 13:
            m = 1
            y += 1


def _safe_filename(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", (name or "").strip())


def _amount_to_text(value) -> str:
    return f"{float(value or 0):.2f} AZN"


def _build_xlsx(title: str, headers: list[str], rows: list[list[str]]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append([title])
    ws.append(headers)
    for row in rows:
        ws.append(row)

    for idx in range(1, len(headers) + 1):
        col = ws.cell(row=2, column=idx).column_letter
        max_len = max(
            len(str(ws.cell(row=r, column=idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col].width = min(max(12, max_len + 2), 40)

    ws.freeze_panes = "A3"
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


ALLOWED_EXPENSE_SORT_KEYS = frozenset({"tip", "tarix", "category", "ad", "mebleg", "sablon", "odenis"})


def _build_admin_expense_table_rows(templates, one_off_expenses, template_expense_by_template_id):
    rows = []
    for t in templates:
        rows.append({"kind": "template", "template": t, "expense": template_expense_by_template_id.get(t.id)})
    for e in one_off_expenses:
        rows.append({"kind": "one_off", "template": None, "expense": e})
    return rows


def _expense_row_stable_tie(row) -> tuple:
    if row["kind"] == "template":
        ex = row["expense"]
        return (0, row["template"].id, ex.id if ex else 0)
    return (1, row["expense"].id, 0)


def _normalize_expense_sort_dt(dt):
    if dt is None:
        return datetime.min.replace(tzinfo=timezone.utc)
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt


def _sort_admin_expense_table_rows(rows: list, sort_col: str, sort_dir: str) -> None:
    if sort_col not in ALLOWED_EXPENSE_SORT_KEYS:
        return
    reverse = sort_dir == "desc"
    tie = _expense_row_stable_tie

    def key_tip(r):
        return ((0 if r["kind"] == "template" else 1,) + tie(r))

    def key_tarix(r):
        if r["kind"] == "template":
            dt = r["template"].created_at
        else:
            dt = r["expense"].created_at
        return (_normalize_expense_sort_dt(dt),) + tie(r)

    def key_category(r):
        if r["kind"] == "template":
            cat = (r["template"].category or "").strip().lower()
        else:
            cat = (r["expense"].category or "").strip().lower()
        return (cat,) + tie(r)

    def key_ad(r):
        if r["kind"] == "template":
            name = (r["template"].name or "").strip().lower()
        else:
            name = (r["expense"].name or "").strip().lower()
        return (name,) + tie(r)

    def key_mebleg(r):
        if r["kind"] == "template":
            t = r["template"]
            ex = r["expense"]
            amt = float(ex.amount) if ex else float(t.default_amount or 0)
        else:
            amt = float(r["expense"].amount or 0)
        return (amt,) + tie(r)

    def key_sablon(r):
        if r["kind"] == "template":
            t = r["template"]
            return (0, 0 if t.is_active else 1, (t.name or "").lower()) + tie(r)
        return (1, 0, "") + tie(r)

    def key_odenis(r):
        if r["kind"] == "template":
            ex = r["expense"]
            if not ex:
                return (2, 0) + tie(r)
            return (0 if ex.is_paid else 1, ex.id) + tie(r)
        e = r["expense"]
        return (0 if e.is_paid else 1, e.id) + tie(r)

    key_fn = {
        "tip": key_tip,
        "tarix": key_tarix,
        "category": key_category,
        "ad": key_ad,
        "mebleg": key_mebleg,
        "sablon": key_sablon,
        "odenis": key_odenis,
    }.get(sort_col)
    if key_fn:
        rows.sort(key=key_fn, reverse=reverse)


def _get_admin_expenses_view_data(period: str):
    templates = ExpenseTemplate.query.order_by(ExpenseTemplate.is_active.desc(), ExpenseTemplate.name.asc()).all()
    expenses = Expense.query.filter_by(period=period).order_by(Expense.created_at.desc()).all()
    template_expense_by_template_id = {e.template_id: e for e in expenses if e.template_id}
    one_off_expenses = [e for e in expenses if not e.template_id]
    total = round(sum(e.amount for e in expenses), 2)
    return {
        "templates": templates,
        "expenses": expenses,
        "one_off_expenses": one_off_expenses,
        "period": period,
        "expenses_total": total,
        "template_expense_by_template_id": template_expense_by_template_id,
        "expense_categories": EXPENSE_CATEGORIES,
    }


def _get_admin_invoices_view_data(selected_period: str, building_id: Optional[int] = None):
    period_rows = (
        db.session.query(Invoice.period)
        .filter(Invoice.period.isnot(None), Invoice.period != "")
        .group_by(Invoice.period)
        .order_by(Invoice.period.desc())
        .all()
    )
    available_periods = [p for (p,) in period_rows]
    if not selected_period and available_periods:
        selected_period = available_periods[0]
    if selected_period and selected_period not in available_periods:
        selected_period = available_periods[0] if available_periods else ""

    apartment_number_sort = db.cast(Apartment.number, db.Integer)
    invoices_query = (
        Invoice.query.options(
            joinedload(Invoice.payments),
            joinedload(Invoice.apartment).joinedload(Apartment.owner),
            joinedload(Invoice.apartment).joinedload(Apartment.building),
        )
        .join(Apartment, Invoice.apartment_id == Apartment.id)
    )
    if selected_period:
        invoices_query = invoices_query.filter(Invoice.period == selected_period)
    if building_id:
        invoices_query = invoices_query.filter(Apartment.building_id == building_id)
    building_null_sort = db.case((Apartment.building_id.is_(None), 1), else_=0)
    invoices = invoices_query.order_by(building_null_sort.asc(), Apartment.building_id.asc(), apartment_number_sort.asc(), Apartment.number.asc()).all()

    dirty = False
    for inv in invoices:
        expected = INVOICE_STATUS_PAID if float(inv.paid_amount or 0) >= float(inv.amount or 0) else "gozlemede"
        if inv.status != expected:
            inv.status = expected
            dirty = True
    if dirty:
        db.session.commit()

    n_inv = len(invoices)
    if n_inv > 0:
        paid_n = sum(1 for inv in invoices if float(inv.paid_amount or 0) >= float(inv.amount or 0))
        period_apartments_paid_pct = round(100.0 * paid_n / float(n_inv), 1)
        period_apartments_late_pct = round(100.0 * (n_inv - paid_n) / float(n_inv), 1)
    else:
        period_apartments_paid_pct = 0.0
        period_apartments_late_pct = 0.0

    prev_period = None
    next_period = None
    if selected_period and selected_period in available_periods:
        idx = available_periods.index(selected_period)
        if idx > 0:
            next_period = available_periods[idx - 1]
        if idx < len(available_periods) - 1:
            prev_period = available_periods[idx + 1]

    return {
        "invoices": invoices,
        "selected_period": selected_period,
        "available_periods": available_periods,
        "prev_period": prev_period,
        "next_period": next_period,
        "building_id": building_id,
        "buildings_all": Building.query.order_by(Building.name.asc()).all(),
        "period_apartments_paid_pct": period_apartments_paid_pct,
        "period_apartments_late_pct": period_apartments_late_pct,
    }


def _get_admin_history_view_data(selected_month: str, building_id: Optional[int] = None):
    month_rows = (
        db.session.query(month_sql_expr(Payment.created_at).label("month"))
        .filter(Payment.status == "confirmed", Payment.created_at.isnot(None))
        .union(
            db.session.query(month_sql_expr(BalanceTopUp.created_at).label("month")).filter(
                BalanceTopUp.created_at.isnot(None)
            ),
            db.session.query(month_sql_expr(Expense.created_at).label("month")).filter(
                Expense.created_at.isnot(None)
            ),
        )
        .all()
    )
    available_months = sorted([m for (m,) in month_rows if m], reverse=True)

    if not selected_month and available_months:
        selected_month = available_months[0]
    if selected_month and selected_month not in available_months:
        selected_month = available_months[0] if available_months else ""

    month_start = None
    month_end = None
    if selected_month:
        year, month = [int(x) for x in selected_month.split("-")]
        month_start = datetime(year, month, 1, tzinfo=timezone.utc)
        month_end = (
            datetime(year + 1, 1, 1, tzinfo=timezone.utc)
            if month == 12
            else datetime(year, month + 1, 1, tzinfo=timezone.utc)
        )

    limit = 500
    payments_query = (
        Payment.query.join(Invoice, Payment.invoice_id == Invoice.id)
        .join(Apartment, Invoice.apartment_id == Apartment.id)
        .filter(Payment.status == "confirmed")
    )
    topups_query = BalanceTopUp.query
    expenses_query = Expense.query
    if month_start and month_end:
        payments_query = payments_query.filter(Payment.created_at >= month_start, Payment.created_at < month_end)
        topups_query = topups_query.filter(BalanceTopUp.created_at >= month_start, BalanceTopUp.created_at < month_end)
        expenses_query = expenses_query.filter(Expense.created_at >= month_start, Expense.created_at < month_end)

    if building_id:
        payments_query = payments_query.filter(Apartment.building_id == building_id)

    confirmed_payments = payments_query.order_by(Payment.created_at.desc()).limit(limit).all()
    topups = topups_query.order_by(BalanceTopUp.created_at.desc()).limit(limit).all()
    expenses = expenses_query.order_by(Expense.created_at.desc()).limit(limit).all()

    events = []
    for p in confirmed_payments:
        inv = p.invoice
        apt = inv.apartment if inv else None
        apt_label = None
        if apt:
            apt_label = f"{apt.building.name} / {apt.number}" if apt.building else apt.number
        events.append(
            {
                "dt": p.created_at,
                "type": "payment",
                "amount": float(p.amount),
                "apartment": apt_label,
                "comment": (p.comment or "").strip() or None,
                "period": inv.period if inv else None,
            }
        )
    for t in topups:
        events.append(
            {
                "dt": t.created_at,
                "type": "topup",
                "amount": float(t.amount),
                "apartment": None,
                "comment": (t.comment or "").strip() or None,
                "period": None,
            }
        )
    for e in expenses:
        events.append(
            {
                "dt": e.created_at,
                "type": "expense",
                "amount": -float(e.amount),
                "apartment": None,
                "comment": (e.name or "").strip() or None,
                "period": e.period,
                "paid": bool(e.is_paid),
            }
        )

    # Сортировка устойчива к смеси naive/aware и к None:
    # naive значения трактуем как UTC, отсутствующие — как -inf.
    _min_dt = datetime.min.replace(tzinfo=timezone.utc)

    def _sort_key(ev):
        dt = ev.get("dt")
        if dt is None:
            return _min_dt
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt

    events.sort(key=_sort_key, reverse=True)
    events = events[:limit]

    prev_month = None
    next_month = None
    if selected_month and selected_month in available_months:
        idx = available_months.index(selected_month)
        if idx > 0:
            next_month = available_months[idx - 1]
        if idx < len(available_months) - 1:
            prev_month = available_months[idx + 1]

    return {
        "events": events,
        "selected_month": selected_month,
        "available_months": available_months,
        "prev_month": prev_month,
        "next_month": next_month,
        "building_id": building_id,
        "buildings_all": Building.query.order_by(Building.name.asc()).all(),
    }


def _get_admin_payments_report_view_data(from_period: str, to_period: str, apartment_id_raw: str, building_id_raw: str = ""):
    apartment_id = int(apartment_id_raw) if apartment_id_raw.isdigit() else None
    building_id = int(building_id_raw) if building_id_raw.isdigit() else None
    periods = list(_iter_months_inclusive(from_period, to_period))
    if len(periods) > 24:
        periods = periods[-24:]
        from_period = periods[0]

    apartment_number_sort = db.cast(Apartment.number, db.Integer)
    building_null_sort = db.case((Apartment.building_id.is_(None), 1), else_=0)
    apartments_query = Apartment.query.order_by(building_null_sort.asc(), Apartment.building_id.asc(), apartment_number_sort.asc(), Apartment.number.asc())
    if apartment_id:
        apartments_query = apartments_query.filter(Apartment.id == apartment_id)
    if building_id:
        apartments_query = apartments_query.filter(Apartment.building_id == building_id)
    apartments = apartments_query.all()

    sums_query = (
        db.session.query(Apartment.id, Invoice.period, db.func.sum(Payment.amount))
        .join(Invoice, Invoice.apartment_id == Apartment.id)
        .join(Payment, Payment.invoice_id == Invoice.id)
        .filter(Payment.status == "confirmed", Invoice.period >= from_period, Invoice.period <= to_period)
    )
    if apartment_id:
        sums_query = sums_query.filter(Apartment.id == apartment_id)
    if building_id:
        sums_query = sums_query.filter(Apartment.building_id == building_id)
    sums = sums_query.group_by(Apartment.id, Invoice.period).all()
    amount_by_apt_period = {(apt_id, period): float(total or 0) for apt_id, period, total in sums}

    rows = []
    for a in apartments:
        row = {
            "apartment": a.number,
            "building": a.building.name if a.building else None,
            "amounts": [],
            "row_total": 0.0,
        }
        for p in periods:
            v = amount_by_apt_period.get((a.id, p), 0.0)
            row["amounts"].append(v)
            row["row_total"] += float(v)
        row["row_total"] = round(row["row_total"], 2)
        rows.append(row)

    col_totals = [round(sum(r["amounts"][idx] for r in rows), 2) for idx in range(len(periods))]
    grand_total = round(sum(col_totals), 2)
    buildings_all = Building.query.order_by(Building.name.asc()).all()
    return {
        "from_period": from_period,
        "to_period": to_period,
        "apartment_id": apartment_id,
        "building_id": building_id,
        "apartments_all": Apartment.query.order_by(Apartment.number).all(),
        "buildings_all": buildings_all,
        "periods": periods,
        "rows": rows,
        "col_totals": col_totals,
        "grand_total": grand_total,
    }


@app.route("/admin/payments-report")
@login_required
@role_required("komendant", "admin")
def admin_payments_report():
    # Payments per apartment per month (confirmed payments).
    today = date.today()
    default_to = today.strftime("%Y-%m")
    default_from = f"{today.year:04d}-{max(1, today.month - 5):02d}"

    from_period = (request.args.get("from_period") or default_from).strip()
    to_period = (request.args.get("to_period") or default_to).strip()
    apartment_id_raw = (request.args.get("apartment_id") or "").strip()
    building_id_raw = (request.args.get("building_id") or "").strip()
    if len(from_period) != 7 or len(to_period) != 7:
        flash("Period duzgun deyil (YYYY-MM).", "danger")
        return redirect(url_for("dashboard"))
    return render_template("admin_payments_report.html", **_get_admin_payments_report_view_data(from_period, to_period, apartment_id_raw, building_id_raw))


@app.route("/admin/export/payments-report/xlsx")
@login_required
@role_required("komendant", "admin")
def export_payments_report():
    today = date.today()
    default_to = today.strftime("%Y-%m")
    default_from = f"{today.year:04d}-{max(1, today.month - 5):02d}"
    from_period = (request.args.get("from_period") or default_from).strip()
    to_period = (request.args.get("to_period") or default_to).strip()
    apartment_id_raw = (request.args.get("apartment_id") or "").strip()
    building_id_raw = (request.args.get("building_id") or "").strip()
    data = _get_admin_payments_report_view_data(from_period, to_period, apartment_id_raw, building_id_raw)

    headers = ["Korpus", "Mənzil"] + data["periods"] + ["Cəmi"]
    rows = []
    for r in data["rows"]:
        rows.append([r["building"] or "", r["apartment"], *[_amount_to_text(v) for v in r["amounts"]], _amount_to_text(r["row_total"])])
    rows.append(["Cəmi", *[_amount_to_text(v) for v in data["col_totals"]], _amount_to_text(data["grand_total"])])
    title = f"Ödənişlər (mənzil/ay): {data['from_period']} - {data['to_period']}"
    filename = _safe_filename(f"payments_report_{data['from_period']}_{data['to_period']}.xlsx")
    payload = _build_xlsx(title, headers, rows)
    return send_file(payload, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/export/expenses/xlsx")
@login_required
@role_required("komendant", "admin")
def export_expenses():
    period = (request.args.get("period") or date.today().strftime("%Y-%m")).strip()
    data = _get_admin_expenses_view_data(period)
    headers = ["Tarix", "Tip", "Kateqoriya", "Ad", "Məbləğ", "Status"]
    rows = [
        [
            e.created_at.strftime("%d.%m.%Y %H:%M") if e.created_at else "",
            "Aylıq" if e.template_id else "Birdəfəlik",
            e.category or "",
            e.name,
            _amount_to_text(e.amount),
            "ödənilib" if e.is_paid else "gözləmədə",
        ]
        for e in data["expenses"]
    ]
    rows.append(["", "", "", "Cəmi", _amount_to_text(data["expenses_total"]), "", ""])
    title = f"Xərclər: {period}"
    filename = _safe_filename(f"expenses_{period}.xlsx")
    payload = _build_xlsx(title, headers, rows)
    return send_file(payload, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/export/invoices/xlsx")
@login_required
@role_required("komendant", "admin")
def export_invoices():
    selected_period = (request.args.get("period", "") or "").strip()
    building_id_raw = (request.args.get("building_id") or "").strip()
    building_id = int(building_id_raw) if building_id_raw.isdigit() else None
    data = _get_admin_invoices_view_data(selected_period, building_id)
    headers = ["Korpus", "Mənzil", "Period", "Hesablanıb", "ödənilib", "Balans", "Status"]
    rows = []
    for i in data["invoices"]:
        balance = float(i.paid_amount or 0) - float(i.amount or 0)
        status = "ödənilib" if float(i.paid_amount or 0) >= float(i.amount or 0) else "ödənilməyib"
        rows.append(
            [
                i.apartment.building.name if i.apartment and i.apartment.building else "",
                i.apartment.number if i.apartment else "",
                i.period,
                _amount_to_text(i.amount),
                _amount_to_text(i.paid_amount),
                _amount_to_text(balance),
                status,
            ]
        )
    title = f"Ödənişlər: {data['selected_period'] or 'bütün periodlar'}"
    filename = _safe_filename(f"invoices_{(data['selected_period'] or 'all')}.xlsx")
    payload = _build_xlsx(title, headers, rows)
    return send_file(payload, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/export/history/xlsx")
@login_required
@role_required("komendant", "admin")
def export_history():
    selected_month = (request.args.get("month") or "").strip()
    building_id_raw = (request.args.get("building_id") or "").strip()
    building_id = int(building_id_raw) if building_id_raw.isdigit() else None
    data = _get_admin_history_view_data(selected_month, building_id)
    headers = ["Tarix", "Növ", "Mənzil", "Period", "Status", "Məbləğ", "Qeyd"]
    rows = []
    for e in data["events"]:
        event_type = "Ödəniş" if e["type"] == "payment" else ("Mədaxil" if e["type"] == "topup" else "Xərc")
        status = "-"
        if e["type"] == "expense":
            status = "ödənilib" if e.get("paid") else "gözləmədə"
        rows.append(
            [
                e["dt"].strftime("%d.%m.%Y %H:%M") if e.get("dt") else "",
                event_type,
                e.get("apartment") or "-",
                e.get("period") or "-",
                status,
                _amount_to_text(e.get("amount")),
                e.get("comment") or "",
            ]
        )
    title = f"Tarixçə: {data['selected_month'] or 'bütün aylar'}"
    filename = _safe_filename(f"history_{(data['selected_month'] or 'all')}.xlsx")
    payload = _build_xlsx(title, headers, rows)
    return send_file(payload, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _render_print_report(title: str, headers: list[str], rows: list[list[str]]):
    return render_template("admin_report_print.html", title=title, headers=headers, rows=rows)


def _render_invoice_print(invoice: Invoice, *, staff_send_channels: bool):
    """Общий helper печати инвойса для admin/komendant и резидента."""
    cfg = get_smtp_config()
    return render_template(
        "invoice_print.html",
        invoice=invoice,
        resident=invoice.apartment.owner,
        cfg=cfg,
        system_name=(cfg.system_name or "").strip() or "eMTK",
        issue_date=utc_to_local(invoice.created_at),
        staff_send_channels=staff_send_channels,
    )


def _render_payment_receipt(payment: Payment, *, staff_send_channels: bool):
    """Общий helper печати квитанции для admin/komendant и резидента."""
    cfg = get_smtp_config()
    return render_template(
        "receipt.html",
        payment=payment,
        cfg=cfg,
        system_name=(cfg.system_name or "").strip() or "eMTK",
        staff_send_channels=staff_send_channels,
    )


@app.route("/admin/print/payments-report")
@login_required
@role_required("komendant", "admin")
def print_payments_report():
    today = date.today()
    default_to = today.strftime("%Y-%m")
    default_from = f"{today.year:04d}-{max(1, today.month - 5):02d}"
    from_period = (request.args.get("from_period") or default_from).strip()
    to_period = (request.args.get("to_period") or default_to).strip()
    apartment_id_raw = (request.args.get("apartment_id") or "").strip()
    building_id_raw = (request.args.get("building_id") or "").strip()
    data = _get_admin_payments_report_view_data(from_period, to_period, apartment_id_raw, building_id_raw)
    headers = ["Korpus", "Mənzil"] + data["periods"] + ["Cəmi"]
    rows = [[r["building"] or "", r["apartment"], *[_amount_to_text(v) for v in r["amounts"]], _amount_to_text(r["row_total"])] for r in data["rows"]]
    rows.append(["", "Cəmi", *[_amount_to_text(v) for v in data["col_totals"]], _amount_to_text(data["grand_total"])])
    return _render_print_report(f"Ödənişlər (mənzil/ay): {data['from_period']} - {data['to_period']}", headers, rows)


@app.route("/admin/print/expenses")
@login_required
@role_required("komendant", "admin")
def print_expenses():
    period = (request.args.get("period") or date.today().strftime("%Y-%m")).strip()
    data = _get_admin_expenses_view_data(period)
    headers = ["Tarix", "Tip", "Kateqoriya", "Ad", "Məbləğ", "Status"]
    rows = [
        [
            e.created_at.strftime("%d.%m.%Y %H:%M") if e.created_at else "",
            "Aylıq" if e.template_id else "Birdəfəlik",
            e.category or "",
            e.name,
            _amount_to_text(e.amount),
            "ödənilib" if e.is_paid else "gözləmədə",
        ]
        for e in data["expenses"]
    ]
    rows.append(["", "", "", "Cəmi", _amount_to_text(data["expenses_total"]), "", ""])
    return _render_print_report(f"Xərclər: {period}", headers, rows)


@app.route("/admin/print/invoices")
@login_required
@role_required("komendant", "admin")
def print_invoices():
    selected_period = (request.args.get("period", "") or "").strip()
    building_id_raw = (request.args.get("building_id") or "").strip()
    building_id = int(building_id_raw) if building_id_raw.isdigit() else None
    data = _get_admin_invoices_view_data(selected_period, building_id)
    headers = ["Korpus", "Mənzil", "Period", "Hesablanıb", "ödənilib", "Balans", "Status"]
    rows = []
    for i in data["invoices"]:
        balance = float(i.paid_amount or 0) - float(i.amount or 0)
        rows.append(
            [
                i.apartment.building.name if i.apartment and i.apartment.building else "",
                i.apartment.number if i.apartment else "",
                i.period,
                _amount_to_text(i.amount),
                _amount_to_text(i.paid_amount),
                _amount_to_text(balance),
                "ödənilib" if float(i.paid_amount or 0) >= float(i.amount or 0) else "ödənilməyib",
            ]
        )
    return _render_print_report(f"Ödənişlər: {data['selected_period'] or 'bütün periodlar'}", headers, rows)


@app.route("/admin/print/history")
@login_required
@role_required("komendant", "admin")
def print_history():
    selected_month = (request.args.get("month") or "").strip()
    building_id_raw = (request.args.get("building_id") or "").strip()
    building_id = int(building_id_raw) if building_id_raw.isdigit() else None
    data = _get_admin_history_view_data(selected_month, building_id)
    headers = ["Tarix", "Növ", "Mənzil", "Period", "Status", "Məbləğ", "Qeyd"]
    rows = []
    for e in data["events"]:
        event_type = "Ödəniş" if e["type"] == "payment" else ("Mədaxil" if e["type"] == "topup" else "Xərc")
        status = "-"
        if e["type"] == "expense":
            status = "ödənilib" if e.get("paid") else "gözləmədə"
        rows.append(
            [
                e["dt"].strftime("%d.%m.%Y %H:%M") if e.get("dt") else "",
                event_type,
                e.get("apartment") or "-",
                e.get("period") or "-",
                status,
                _amount_to_text(e.get("amount")),
                e.get("comment") or "",
            ]
        )
    return _render_print_report(f"Tarixçə: {data['selected_month'] or 'bütün aylar'}", headers, rows)


@app.route("/resident/receipt/<int:payment_id>")
@login_required
@role_required("resident")
def resident_receipt(payment_id):
    payment = Payment.query.get_or_404(payment_id)
    user = current_user()
    if payment.invoice.apartment.owner_user_id != user.id:
        flash("Qəbzə giriş mümkün deyil.", "danger")
        return redirect(url_for("dashboard"))
    return _render_payment_receipt(payment, staff_send_channels=False)


@app.route("/resident/invoice/<int:invoice_id>")
@login_required
@role_required("resident")
def resident_invoice_print(invoice_id):
    """Печатная версия инвойса для резидента — используется в WA-ссылках."""
    invoice = Invoice.query.get_or_404(invoice_id)
    user = current_user()
    if invoice.apartment.owner_user_id != user.id:
        flash("Hesab-fakturaya giriş mümkün deyil.", "danger")
        return redirect(url_for("dashboard"))
    return _render_invoice_print(invoice, staff_send_channels=False)


@app.route("/admin/content", methods=["GET", "POST"])
@login_required
@role_required("komendant", "admin")
def admin_content():
    content_type = request.args.get("type", "work")
    if request.method == "POST":
        form_type = request.form.get("form_type")
        if form_type == "work":
            before_photo = save_uploaded_image(request.files.get("before_photo_file"))
            after_photo = save_uploaded_image(request.files.get("after_photo_file"))
            db.session.add(
                WorkLog(
                    title=request.form["title"].strip(),
                    description=request.form["description"].strip(),
                    before_photo_url=before_photo,
                    after_photo_url=after_photo,
                )
            )
            audit("Is jurnalina qeyd elave edildi")
            flash("Is elave edildi.", "success")
            sysname = (get_smtp_config().system_name or "").strip() or "eMTK"
            notify_residents(f"{sysname}: Yeni isler", f"Yeni is elave edildi: {request.form['title'].strip()}")
        elif form_type == "announcement":
            title = request.form.get("title", "").strip()
            text = request.form.get("text", "").strip()
            if not title or not text:
                flash("Başlıq və mətn vacibdir.", "danger")
                return redirect(url_for("admin_content", type="announcement"))
            db.session.add(Announcement(title=title, text=text))
            audit("Yeni elan elave edildi")
            flash("Elan elave edildi.", "success")
            sysname = (get_smtp_config().system_name or "").strip() or "eMTK"
            notify_residents(f"{sysname}: Yeni elan", f"Yeni elan: {title}")
        db.session.commit()
        return redirect(url_for("admin_content", type=content_type))

    works = WorkLog.query.order_by(WorkLog.created_at.desc()).all()
    announcements = Announcement.query.order_by(Announcement.created_at.desc()).all()
    residents_for_wa = (
        User.query.options(joinedload(User.apartments).joinedload(Apartment.building))
        .filter_by(role="resident")
        .order_by(User.full_name.asc())
        .all()
    )
    return render_template(
        "admin_content.html",
        works=works,
        announcements=announcements,
        content_type=content_type,
        residents_for_wa=residents_for_wa,
    )


@app.route("/admin/announcements/update/<int:announcement_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def update_announcement(announcement_id):
    announcement = Announcement.query.get_or_404(announcement_id)
    announcement.title = request.form["title"].strip()
    announcement.text = request.form["text"].strip()
    db.session.commit()
    audit(f"Elan yenilendi #{announcement_id}")
    flash("Elan yenilendi.", "success")
    return redirect(url_for("admin_content", type="announcement"))


@app.route("/admin/announcements/delete/<int:announcement_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def delete_announcement(announcement_id):
    announcement = Announcement.query.get_or_404(announcement_id)
    db.session.delete(announcement)
    db.session.commit()
    audit(f"Elan silindi #{announcement_id}")
    flash("Elan silindi.", "success")
    return redirect(url_for("admin_content", type="announcement"))


@app.route("/admin/worklogs/update/<int:worklog_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def update_worklog(worklog_id):
    w = WorkLog.query.get_or_404(worklog_id)
    w.title = request.form["title"].strip()
    w.description = request.form["description"].strip()

    before_photo = save_uploaded_image(request.files.get("before_photo_file"))
    after_photo = save_uploaded_image(request.files.get("after_photo_file"))
    if before_photo:
        w.before_photo_url = before_photo
    if after_photo:
        w.after_photo_url = after_photo

    db.session.commit()
    audit(f"Is yenilendi #{worklog_id}")
    flash("Is yenilendi.", "success")
    return redirect(url_for("admin_content", type="work"))


@app.route("/admin/worklogs/delete/<int:worklog_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def delete_worklog(worklog_id):
    w = WorkLog.query.get_or_404(worklog_id)
    db.session.delete(w)
    db.session.commit()
    audit(f"Is silindi #{worklog_id}")
    flash("Is silindi.", "success")
    return redirect(url_for("admin_content", type="work"))


@app.route("/admin/content/whatsapp-broadcast", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def admin_whatsapp_broadcast():
    wa_cfg = get_whatsapp_config()
    if not (wa_cfg.enabled and wa_cfg.api_url and wa_cfg.api_key and wa_cfg.instance):
        flash("WhatsApp inteqrasiyası deaktivdir və ya ayarlar tam deyil.", "warning")
        return redirect(url_for("admin_content"))
    scope = (request.form.get("scope") or "bulk").strip().lower()
    message = (request.form.get("message") or "").strip()
    if not message:
        flash("Mesaj mətni boş ola bilməz.", "danger")
        return redirect(url_for("admin_content"))
    if scope == "target":
        raw_ids = request.form.getlist("user_ids")
        uid_list: list[int] = []
        for x in raw_ids:
            try:
                uid_list.append(int(x))
            except (TypeError, ValueError):
                continue
        if not uid_list:
            flash("Ən azı bir sakin seçin.", "danger")
            return redirect(url_for("admin_content"))
        users = User.query.filter(User.id.in_(uid_list), User.role == "resident").all()
    else:
        users = User.query.filter_by(role="resident").all()
    enq, skip = wa_broadcast_enqueue_personal(users, message)
    try:
        wa_queue_drain_once()
    except Exception:
        pass
    audit(f"WhatsApp mesaj növbəsi: scope={scope}, queued={enq}, no_phone={skip}")
    if enq:
        flash(
            f"{enq} alıcı növbəyə əlavə edildi. Telefonu olmayan / təkrar nömrə: {skip}. "
            f"Limit: {wa_cfg.bulk_limit} mesaj / {wa_cfg.bulk_window_sec} san — qalanlar avtomatik göndəriləcək.",
            "success",
        )
    else:
        flash("Heç bir alıcı növbəyə düşmədi (telefon yoxdur və ya mətn boşdur).", "warning")
    return redirect(url_for("admin_content"))


@app.route("/admin/content/whatsapp-from-item", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def admin_whatsapp_from_item():
    wa_cfg = get_whatsapp_config()
    if not (wa_cfg.enabled and wa_cfg.api_url and wa_cfg.api_key and wa_cfg.instance):
        flash("WhatsApp inteqrasiyası deaktivdir və ya ayarlar tam deyil.", "warning")
        return redirect(url_for("admin_content"))
    kind = (request.form.get("kind") or "").strip().lower()
    scope = (request.form.get("scope") or "bulk").strip().lower()
    try:
        item_id = int(request.form.get("item_id") or "0")
    except (TypeError, ValueError):
        flash("Yanlış identifikator.", "danger")
        return redirect(url_for("admin_content"))
    smtp = get_smtp_config()
    if kind == "work":
        w = WorkLog.query.get_or_404(item_id)
        text = build_whatsapp_content_broadcast_text("is", w.title, w.description, smtp)
        redir_type = "work"
    elif kind == "announcement":
        a = Announcement.query.get_or_404(item_id)
        text = build_whatsapp_content_broadcast_text("elan", a.title, a.text, smtp)
        redir_type = "announcement"
    else:
        flash("Yanlış növ.", "danger")
        return redirect(url_for("admin_content"))
    if scope == "target":
        raw_ids = request.form.getlist("user_ids")
        uid_list: list[int] = []
        for x in raw_ids:
            try:
                uid_list.append(int(x))
            except (TypeError, ValueError):
                continue
        if not uid_list:
            flash("Ən azı bir sakin seçin.", "danger")
            return redirect(url_for("admin_content", type=redir_type))
        users = User.query.filter(User.id.in_(uid_list), User.role == "resident").all()
    else:
        users = User.query.filter_by(role="resident").all()
    enq, skip = wa_broadcast_enqueue_personal(users, text)
    try:
        wa_queue_drain_once()
    except Exception:
        pass
    audit(f"WhatsApp elan/is #{item_id} ({kind}): queued={enq}, scope={scope}")
    if enq:
        flash(
            f"{enq} alıcı növbəyə əlavə edildi. Telefonu olmayan: {skip}.",
            "success",
        )
    else:
        flash("Heç bir alıcı növbəyə düşmədi.", "warning")
    return redirect(url_for("admin_content", type=redir_type))


@app.route("/admin/users")
@login_required
@role_required("komendant", "admin")
def admin_users():
    sort = (request.args.get("sort") or "name").strip().lower()
    sort_dir = (request.args.get("dir") or "asc").strip().lower()
    if sort_dir not in ("asc", "desc"):
        sort_dir = "asc"
    allowed_sorts = ("id", "name", "apartment", "phone", "email", "role", "wa")
    if sort not in allowed_sorts:
        sort = "name"
        sort_dir = "asc"

    users = User.query.options(joinedload(User.apartments).joinedload(Apartment.building)).all()
    for u in users:
        u.apartments.sort(key=_apartment_row_sort_key)

    reverse = sort_dir == "desc"

    if sort == "id":
        users.sort(key=lambda u: u.id, reverse=reverse)
    elif sort == "name":
        users.sort(key=lambda u: ((u.full_name or "").lower(), u.id), reverse=reverse)
    elif sort == "phone":
        users.sort(key=lambda u: ((u.phone or "").lower(), u.id), reverse=reverse)
    elif sort == "email":
        users.sort(key=lambda u: ((u.email or "").lower(), u.id), reverse=reverse)
    elif sort == "role":
        users.sort(key=lambda u: ((u.role or "").lower(), u.id), reverse=reverse)
    elif sort == "wa":
        users.sort(key=lambda u: ((1 if u.whatsapp_connected else 0), u.id), reverse=reverse)
    elif sort == "apartment":
        def _user_apt_sort_key(u: User):
            if not u.apartments:
                return (1, "", 0, 10**9, "", u.id)
            b_name, b_id, n, tail = _apartment_row_sort_key(u.apartments[0])
            return (0, b_name, b_id, n, tail, u.id)

        users.sort(key=_user_apt_sort_key, reverse=reverse)

    users_sort_next = {}
    for c in allowed_sorts:
        if sort != c:
            users_sort_next[c] = {"sort": c, "dir": "asc"}
        else:
            users_sort_next[c] = {"sort": c, "dir": ("desc" if sort_dir == "asc" else "asc")}

    return render_template(
        "admin_users.html",
        users=users,
        users_sort=sort,
        users_sort_dir=sort_dir,
        users_sort_next=users_sort_next,
    )


@app.route("/admin/users/create", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def admin_user_create():
    full_name = (request.form.get("full_name") or "").strip()
    phone = (request.form.get("phone") or "").strip() or None
    email = (request.form.get("email") or "").strip().lower()
    role = (request.form.get("role") or "resident").strip()
    password = request.form.get("password") or ""

    if not full_name or not email or not password:
        flash("Zorunlu sahələr boş ola bilməz.", "danger")
        return redirect(url_for("admin_users"))
    if phone and not normalize_az_phone(phone):
        flash("Telefon formatı +994XXXXXXXXX olmalıdır.", "danger")
        return redirect(url_for("admin_users"))
    if role == "commandant":
        role = "komendant"
    if role == "superadmin":
        role = "admin"
    if role not in ("resident", "komendant", "admin"):
        flash("Rol duzgun deyil.", "danger")
        return redirect(url_for("admin_users"))
    me = current_user()
    if role == "admin" and (not me or me.role != "admin"):
        flash("Yalnız admin admin yarada bilər.", "danger")
        return redirect(url_for("admin_users"))
    if User.query.filter_by(email=email).first():
        flash("Bu email ile artiq istifadəçi var.", "warning")
        return redirect(url_for("admin_users"))

    db.session.add(
        User(
            full_name=full_name,
            phone=phone,
            email=email,
            password_hash=generate_password_hash(password),
            role=role,
        )
    )
    db.session.commit()
    audit(f"İstifadəçi yaradıldı {email} ({role})")
    flash("İstifadəçi yaradıldı.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/update/<int:user_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def admin_user_update(user_id):
    target = User.query.get_or_404(user_id)

    full_name = (request.form.get("full_name") or "").strip()
    phone = (request.form.get("phone") or "").strip() or None
    email = (request.form.get("email") or "").strip().lower()
    role = (request.form.get("role") or target.role).strip()
    new_password = (request.form.get("password") or "").strip()

    if not full_name or not email:
        flash("Ad və email boş ola bilməz.", "danger")
        return redirect(url_for("admin_users"))
    if phone and not normalize_az_phone(phone):
        flash("Telefon formatı +994XXXXXXXXX olmalıdır.", "danger")
        return redirect(url_for("admin_users"))
    if role == "commandant":
        role = "komendant"
    if role == "superadmin":
        role = "admin"
    if role not in ("resident", "komendant", "admin"):
        flash("Rol duzgun deyil.", "danger")
        return redirect(url_for("admin_users"))
    me = current_user()
    if (not me or me.role != "admin") and (target.role == "admin" or role == "admin"):
        flash("Admin hesablarını yalnız admin idarə edə bilər.", "danger")
        return redirect(url_for("admin_users"))
    duplicate = User.query.filter(User.email == email, User.id != target.id).first()
    if duplicate:
        flash("Bu email başqa istifadəçidə var.", "warning")
        return redirect(url_for("admin_users"))

    # Prevent self-demoting to resident (locks admin out).
    if me and me.id == target.id and target.role in ("komendant", "admin") and role == "resident":
        flash("Öz rolunuzu resident edə bilməzsiniz.", "warning")
        return redirect(url_for("admin_users"))

    target.full_name = full_name
    target.phone = phone
    target.email = email
    target.role = role
    if new_password:
        target.password_hash = generate_password_hash(new_password)

    db.session.commit()
    audit(f"İstifadəçi yeniləndi #{target.id} {target.email} ({target.role})")
    flash("İstifadəçi yeniləndi.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/delete/<int:user_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def admin_user_delete(user_id):
    target = User.query.get_or_404(user_id)
    me = current_user()
    if me and me.id == target.id:
        flash("Özünüzü silə bilməzsiniz.", "warning")
        return redirect(url_for("admin_users"))
    if me and me.role != "admin" and target.role == "admin":
        flash("Admin hesabını yalnız admin silə bilər.", "danger")
        return redirect(url_for("admin_users"))

    has_apartments = Apartment.query.filter_by(owner_user_id=target.id).first() is not None
    referenced = (
        Payment.query.filter_by(reviewer_user_id=target.id).first() is not None
        or BalanceTopUp.query.filter_by(created_by_user_id=target.id).first() is not None
        or Expense.query.filter_by(created_by_user_id=target.id).first() is not None
        or AuditLog.query.filter_by(actor_user_id=target.id).first() is not None
    )
    if has_apartments or referenced:
        flash("İstifadəçini silmək olmur: bağlı məlumatlar var.", "warning")
        return redirect(url_for("admin_users"))

    email = target.email
    db.session.delete(target)
    db.session.commit()
    audit(f"İstifadəçi silindi {email}")
    flash("İstifadəçi silindi.", "success")
    return redirect(url_for("admin_users"))


@app.route("/polls", methods=["GET", "POST"])
@login_required
def polls():
    flash("Sorğular və səsvermələr funksiyası deaktiv edilib.", "info")
    return redirect(url_for("dashboard"))


def _redirect_after_invoice_staff_send(invoice: Invoice):
    """Return to print view when sending from invoice_print.html (stay_on_print=1)."""
    if request.form.get("stay_on_print") == "1":
        return redirect(url_for("print_invoice", invoice_id=invoice.id))
    return redirect(url_for("admin_invoices"))


def _redirect_after_payment_receipt_staff_send(payment: Payment):
    if request.form.get("stay_on_print") == "1":
        return redirect(url_for("print_payment_receipt", payment_id=payment.id))
    period = (request.form.get("period") or "").strip()
    bid = (request.form.get("building_id") or "").strip()
    return redirect(url_for("admin_invoices", period=period or payment.invoice.period, building_id=bid or None))


@app.route("/admin/invoices/send/<int:invoice_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def send_invoice_email(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    resident = invoice.apartment.owner
    if not resident.email:
        flash("Sakinin email adresi yoxdur.", "warning")
        return _redirect_after_invoice_staff_send(invoice)
    cfg = get_smtp_config()
    subject, body, html_body = build_invoice_email(invoice, resident, cfg)
    if send_email(subject, body, [resident.email], html_body=html_body):
        flash("Hesab-faktura email ilə göndərildi.", "success")
    else:
        flash("Email gonderilemedi. SMTP ayarlarini yoxlayin.", "danger")
    return _redirect_after_invoice_staff_send(invoice)


@app.route("/admin/payments/send-email/<int:payment_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def send_payment_receipt_email(payment_id):
    payment = Payment.query.get_or_404(payment_id)
    invoice = payment.invoice
    resident = invoice.apartment.owner if invoice and invoice.apartment else None
    if not resident or not resident.email:
        flash("Sakinin email adresi yoxdur.", "warning")
        return _redirect_after_payment_receipt_staff_send(payment)
    cfg = get_smtp_config()
    subject, body, html_body = build_receipt_email(payment, cfg)
    if send_email(subject, body, [resident.email], html_body=html_body):
        flash("Qəbz email ilə göndərildi.", "success")
    else:
        flash("Email gonderilemedi. SMTP ayarlarini yoxlayin.", "danger")
    return _redirect_after_payment_receipt_staff_send(payment)


@app.route("/admin/payments/whatsapp/<int:payment_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def send_payment_receipt_whatsapp(payment_id):
    payment = Payment.query.get_or_404(payment_id)
    if (payment.status or "") != "confirmed":
        flash("Yalnız təsdiqlənmiş ödəniş üçün qəbz göndərilə bilər.", "warning")
        return _redirect_after_payment_receipt_staff_send(payment)
    invoice = payment.invoice
    resident = invoice.apartment.owner if invoice and invoice.apartment else None
    if not resident or not resident.phone:
        flash("Sakinin telefon nömrəsi yoxdur.", "warning")
        return _redirect_after_payment_receipt_staff_send(payment)
    wa_cfg = get_whatsapp_config()
    if not (wa_cfg.enabled and wa_cfg.api_url and wa_cfg.api_key and wa_cfg.instance):
        flash("WhatsApp inteqrasiyası deaktiv və ya ayarları tam deyil.", "warning")
        return _redirect_after_payment_receipt_staff_send(payment)
    smtp_cfg = get_smtp_config()
    text = build_whatsapp_receipt_text(payment, smtp_cfg)
    ok, err = wa_send_text(resident.phone, text)
    if ok:
        audit(f"Ödəniş qəbzi #{payment.id} WhatsApp ilə göndərildi ({resident.phone})")
        flash("Qəbz WhatsApp ilə göndərildi.", "success")
    else:
        flash(f"WhatsApp göndərilmədi: {err}", "danger")
    return _redirect_after_payment_receipt_staff_send(payment)


@app.route("/admin/invoices/whatsapp/<int:invoice_id>", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def send_invoice_whatsapp(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    resident = invoice.apartment.owner
    if not resident or not resident.phone:
        flash("Sakinin telefon nömrəsi yoxdur.", "warning")
        if request.form.get("stay_on_print") == "1":
            return redirect(url_for("print_invoice", invoice_id=invoice.id))
        return redirect(url_for("admin_invoices"))
    wa_cfg = get_whatsapp_config()
    if not (wa_cfg.enabled and wa_cfg.api_url and wa_cfg.api_key and wa_cfg.instance):
        flash("WhatsApp inteqrasiyası deaktiv və ya ayarları tam deyil.", "warning")
        if request.form.get("stay_on_print") == "1":
            return redirect(url_for("print_invoice", invoice_id=invoice.id))
        return redirect(url_for("admin_invoices"))
    smtp_cfg = get_smtp_config()
    text = build_whatsapp_invoice_text(invoice, resident, smtp_cfg)
    ok, err = wa_send_text(resident.phone, text)
    if ok:
        audit(f"Hesab-faktura #{invoice.id} WhatsApp ilə göndərildi ({resident.phone})")
        flash("Hesab-faktura WhatsApp ilə göndərildi.", "success")
    else:
        flash(f"WhatsApp göndərilmədi: {err}", "danger")
    if request.form.get("stay_on_print") == "1":
        return redirect(url_for("print_invoice", invoice_id=invoice.id))
    return redirect(url_for("admin_invoices", period=invoice.period))


@app.route("/admin/invoices/whatsapp/bulk", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def send_invoices_whatsapp_bulk():
    wa_cfg = get_whatsapp_config()
    if not wa_cfg.enabled:
        flash("WhatsApp inteqrasiyası deaktiv edilib.", "warning")
        return redirect(url_for("admin_invoices"))

    period = (request.form.get("period") or "").strip()
    building_id = request.form.get("building_id")
    try:
        building_id = int(building_id) if building_id else None
    except ValueError:
        building_id = None

    q = Invoice.query.join(Apartment, Invoice.apartment_id == Apartment.id)
    if period:
        q = q.filter(Invoice.period == period)
    if building_id:
        q = q.filter(Apartment.building_id == building_id)
    invoices = q.all()

    smtp_cfg = get_smtp_config()
    queued = 0
    skipped = 0
    for inv in invoices:
        resident = inv.apartment.owner
        if not resident or not resident.phone:
            skipped += 1
            continue
        text = build_whatsapp_invoice_text(inv, resident, smtp_cfg)
        if wa_queue_enqueue(resident.phone, text, invoice_id=inv.id, user_id=resident.id):
            queued += 1
        else:
            skipped += 1

    audit(f"Kütləvi WhatsApp: {queued} mesaj növbəyə əlavə edildi (period={period or 'hamısı'})")
    flash(
        f"Növbəyə əlavə edildi: {queued}. Telefonu olmayanlar: {skipped}. "
        f"Limit: {wa_cfg.bulk_limit} mesaj / {wa_cfg.bulk_window_sec // 60} dəq.",
        "success",
    )
    return redirect(url_for("admin_invoices", period=period, building_id=building_id or ""))


@app.route("/admin/invoices/email/bulk", methods=["POST"])
@login_required
@role_required("komendant", "admin")
def send_invoices_email_bulk():
    period = (request.form.get("period") or "").strip()
    building_id_raw = request.form.get("building_id")
    try:
        building_id = int(building_id_raw) if building_id_raw else None
    except ValueError:
        building_id = None

    smtp_cfg = get_smtp_config()
    if not smtp_cfg.host or not smtp_cfg.sender_email:
        flash("SMTP ayarları tam deyil (host və ya göndərən email). Admin parametrlərini yoxlayın.", "danger")
        return redirect(url_for("admin_invoices", period=period, building_id=building_id or ""))

    q = Invoice.query.join(Apartment, Invoice.apartment_id == Apartment.id)
    if period:
        q = q.filter(Invoice.period == period)
    if building_id:
        q = q.filter(Apartment.building_id == building_id)
    invoices = q.all()

    sent = 0
    failed = 0
    skipped = 0
    for inv in invoices:
        resident = inv.apartment.owner
        if not resident or not resident.email:
            skipped += 1
            continue
        subject, body, html_body = build_invoice_email(inv, resident, smtp_cfg)
        if send_email(subject, body, [resident.email], html_body=html_body):
            sent += 1
        else:
            failed += 1

    audit(
        f"Kütləvi email (hesab-faktura): göndərildi={sent}, uğursuz={failed}, email yox={skipped} "
        f"(period={period or 'hamısı'})"
    )
    if failed and not sent:
        flash(
            f"Email göndərilmədi (SMTP xətası və ya bütün cəhdlər uğursuz). "
            f"Emaili olmayanlar: {skipped}.",
            "danger",
        )
    elif failed:
        flash(
            f"Göndərildi: {sent}. Uğursuz: {failed}. Email ünvanı olmayanlar: {skipped}.",
            "warning",
        )
    else:
        flash(
            f"Hesab-faktura email ilə göndərildi: {sent}. Email ünvanı olmayanlar: {skipped}.",
            "success",
        )
    return redirect(url_for("admin_invoices", period=period, building_id=building_id or ""))


@app.route("/admin/invoices/print/<int:invoice_id>")
@login_required
@role_required("komendant", "admin")
def print_invoice(invoice_id):
    invoice = Invoice.query.options(joinedload(Invoice.apartment).joinedload(Apartment.owner)).get_or_404(invoice_id)
    return _render_invoice_print(invoice, staff_send_channels=True)


@app.route("/admin/payments/print/<int:payment_id>")
@login_required
@role_required("komendant", "admin")
def print_payment_receipt(payment_id):
    payment = (
        Payment.query.options(
            joinedload(Payment.invoice).joinedload(Invoice.apartment).joinedload(Apartment.owner)
        )
        .filter_by(id=payment_id)
        .first_or_404()
    )
    return _render_payment_receipt(payment, staff_send_channels=True)


@app.route("/admin/settings", methods=["GET", "POST"])
@login_required
@role_required("admin", "komendant")
def admin_settings():
    cfg = get_smtp_config()
    if request.method == "POST":
        form_type = request.form.get("form_type", "save_smtp")
        if form_type == "save_system":
            cfg.system_name = request.form.get("system_name", "").strip() or None
            cfg.house_address = request.form.get("house_address", "").strip() or None
            cfg.commandant_name = request.form.get("commandant_name", "").strip() or None
            cfg.contact_phone = request.form.get("contact_phone", "").strip() or None
            wg_raw = (request.form.get("whatsapp_group_url", "") or "").strip()
            if wg_raw and not wg_raw.lower().startswith(("http://", "https://")):
                wg_raw = "https://" + wg_raw
            cfg.whatsapp_group_url = wg_raw[:512] or None
            # Portal URL меняет только суперадмин (сверяем по БД, не по сессии).
            acting_user = current_user()
            if acting_user and acting_user.role == "admin":
                portal_raw = (request.form.get("portal_url", "") or "").strip()
                if portal_raw and not portal_raw.lower().startswith(("http://", "https://")):
                    portal_raw = "https://" + portal_raw
                cfg.portal_url = portal_raw.rstrip("/") or None
            db.session.commit()
            audit("Sistem ayarlari yenilendi")
            flash("Sistem ayarlari saxlanildi.", "success")
        elif form_type == "save_smtp":
            cfg.host = request.form.get("host", "").strip() or None
            cfg.port = int(request.form.get("port", "587"))
            cfg.username = request.form.get("username", "").strip() or None
            new_password = request.form.get("password", "").strip()
            if new_password:
                cfg.password = new_password
            cfg.sender_email = request.form.get("sender_email", "").strip() or None
            cfg.use_tls = request.form.get("use_tls") == "on"
            db.session.commit()
            audit("SMTP ayarlari yenilendi")
            flash("SMTP ayarlari saxlanildi.", "success")
        elif form_type == "test_email":
            to_email = request.form.get("test_email", "").strip()
            sysname = (cfg.system_name or "").strip() or "eMTK"
            if send_email(f"{sysname} SMTP test", "Bu test mesajidir.", [to_email]):
                flash("Test email ugurla gonderildi.", "success")
            else:
                flash("Test email gonderilmedi. SMTP ayarlarini yoxlayin.", "danger")
        elif form_type == "save_whatsapp":
            wa_cfg = get_whatsapp_config()
            wa_cfg.enabled = request.form.get("wa_enabled") == "on"
            wa_cfg.api_url = request.form.get("wa_api_url", "").strip() or None
            new_key = request.form.get("wa_api_key", "").strip()
            if new_key:
                wa_cfg.api_key = new_key
            wa_cfg.instance = request.form.get("wa_instance", "").strip() or None
            wa_cfg.service_number = request.form.get("wa_service_number", "").strip() or None
            try:
                wa_cfg.bulk_limit = max(1, int(request.form.get("wa_bulk_limit", "10") or "10"))
            except ValueError:
                wa_cfg.bulk_limit = 10
            try:
                wa_cfg.bulk_window_sec = max(30, int(request.form.get("wa_bulk_window_sec", "300") or "300"))
            except ValueError:
                wa_cfg.bulk_window_sec = 300
            db.session.commit()
            audit("WhatsApp ayarlari yenilendi")
            flash("WhatsApp ayarlari saxlanildi.", "success")
        elif form_type == "test_whatsapp":
            to_phone = request.form.get("test_phone", "").strip()
            sysname = (cfg.system_name or "").strip() or "eMTK"
            ok, err = wa_send_text(to_phone, f"{sysname} — WhatsApp test mesajı.")
            if ok:
                flash("Test WhatsApp mesajı göndərildi.", "success")
            else:
                flash(f"WhatsApp test uğursuz: {err}", "danger")
        elif form_type == "add_apartment_preset":
            name = (request.form.get("name", "") or "").strip()
            rooms_raw = (request.form.get("rooms", "") or "").strip()
            area_raw = (request.form.get("area", "") or "").strip()
            if not name or not rooms_raw or not area_raw:
                flash("Preset üçün ad, otaq və sahə vacibdir.", "danger")
                return redirect(url_for("admin_settings"))
            try:
                rooms = _parse_int_field(rooms_raw, min_value=1, max_value=999)
                area = _parse_int_field(area_raw, min_value=1, max_value=9999)
            except ValueError:
                flash("Preset üçün otaq 1-999, sahə isə 1-9999 aralığında tam ədəd olmalıdır.", "danger")
                return redirect(url_for("admin_settings"))
            if rooms is None or area is None:
                flash("Otaq və sahə sıfırdan böyük olmalıdır.", "danger")
                return redirect(url_for("admin_settings"))
            db.session.add(ApartmentPreset(name=name, rooms=rooms, area=area))
            db.session.commit()
            audit(f"Mənzil preset əlavə edildi: {name} ({rooms} otaq, {area} m2)")
            flash("Mənzil preset əlavə edildi.", "success")
        elif form_type == "update_apartment_preset":
            preset_id = int(request.form["preset_id"])
            preset = ApartmentPreset.query.get_or_404(preset_id)
            name = (request.form.get("name", "") or "").strip()
            rooms_raw = (request.form.get("rooms", "") or "").strip()
            area_raw = (request.form.get("area", "") or "").strip()
            try:
                rooms = _parse_int_field(rooms_raw, min_value=1, max_value=999)
                area = _parse_int_field(area_raw, min_value=1, max_value=9999)
            except ValueError:
                flash("Preset məlumatları düzgün deyil: otaq 1-999, sahə 1-9999 olmalıdır.", "danger")
                return redirect(url_for("admin_settings"))
            if not name or rooms is None or area is None:
                flash("Preset məlumatları düzgün deyil.", "danger")
                return redirect(url_for("admin_settings"))
            preset.name = name
            preset.rooms = rooms
            preset.area = area
            db.session.commit()
            audit(f"Mənzil preset yeniləndi #{preset_id}")
            flash("Mənzil preset yeniləndi.", "success")
        elif form_type == "delete_apartment_preset":
            preset_id = int(request.form["preset_id"])
            preset = ApartmentPreset.query.get_or_404(preset_id)
            preset_name = preset.name
            db.session.delete(preset)
            db.session.commit()
            audit(f"Mənzil preset silindi #{preset_id}: {preset_name}")
            flash("Mənzil preset silindi.", "success")
        elif form_type == "add_building":
            name = (request.form.get("building_name", "") or "").strip()
            address = (request.form.get("building_address", "") or "").strip() or None
            if not name:
                flash("Korpus adı vacibdir.", "danger")
                return redirect(url_for("admin_settings"))
            if Building.query.filter_by(name=name).first():
                flash("Bu adda korpus artıq mövcuddur.", "warning")
                return redirect(url_for("admin_settings"))
            db.session.add(Building(name=name, address=address))
            db.session.commit()
            audit(f"Korpus əlavə edildi: {name}")
            flash("Korpus əlavə edildi.", "success")
        elif form_type == "update_building":
            building_id = int(request.form["building_id"])
            building = Building.query.get_or_404(building_id)
            name = (request.form.get("building_name", "") or "").strip()
            address = (request.form.get("building_address", "") or "").strip() or None
            if not name:
                flash("Korpus adı vacibdir.", "danger")
                return redirect(url_for("admin_settings"))
            duplicate = Building.query.filter(Building.name == name, Building.id != building_id).first()
            if duplicate:
                flash("Bu adda korpus artıq mövcuddur.", "warning")
                return redirect(url_for("admin_settings"))
            building.name = name
            building.address = address
            db.session.commit()
            audit(f"Korpus yeniləndi #{building_id}: {name}")
            flash("Korpus yeniləndi.", "success")
        elif form_type == "delete_building":
            building_id = int(request.form["building_id"])
            building = Building.query.get_or_404(building_id)
            if building.apartments:
                flash("Korpusu silmək olmur: bağlı mənzillər var.", "warning")
                return redirect(url_for("admin_settings"))
            building_name = building.name
            db.session.delete(building)
            db.session.commit()
            audit(f"Korpus silindi #{building_id}: {building_name}")
            flash("Korpus silindi.", "success")
        return redirect(url_for("admin_settings"))
    apartment_presets = ApartmentPreset.query.order_by(ApartmentPreset.rooms.asc(), ApartmentPreset.area.asc()).all()
    buildings = Building.query.order_by(Building.name.asc()).all()
    db_path = sqlite_main_database_path()
    database_backup_supported = db_path is not None and db_path.is_file()
    return render_template(
        "admin_settings.html",
        cfg=cfg,
        wa_cfg=get_whatsapp_config(),
        apartment_presets=apartment_presets,
        buildings=buildings,
        database_backup_supported=database_backup_supported,
    )


@app.route("/admin/settings/database-export")
@login_required
@role_required("admin", "komendant")
def admin_database_export():
    db_path = sqlite_main_database_path()
    if not db_path or not db_path.is_file():
        flash("Verilənlər bazasının ixracı yalnız fayl əsaslı SQLite üçün mövcuddur.", "warning")
        return redirect(url_for("admin_settings"))
    tmp_fd, tmp_name = tempfile.mkstemp(prefix="emtk_export_", suffix=".db")
    os.close(tmp_fd)
    tmp_path = Path(tmp_name)
    try:
        src = sqlite3.connect(str(db_path), timeout=120)
        try:
            dst = sqlite3.connect(str(tmp_path))
            try:
                with dst:
                    src.backup(dst)
            finally:
                dst.close()
        finally:
            src.close()
    except (sqlite3.Error, OSError) as exc:
        tmp_path.unlink(missing_ok=True)
        flash(f"Yedək yaradılmadı: {exc}", "danger")
        return redirect(url_for("admin_settings"))

    stamp = utc_to_local(datetime.now(timezone.utc)).strftime("%Y%m%d_%H%M%S")
    download_name = f"emtk_backup_{stamp}.db"

    @after_this_request
    def _cleanup_export_temp(response):
        try:
            tmp_path.unlink(missing_ok=True)
        except OSError:
            pass
        return response

    audit("Verilənlər bazası ixrac edildi (SQLite .db)")
    return send_file(
        tmp_path,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.sqlite3",
    )


@app.route("/admin/settings/database-import", methods=["POST"])
@login_required
@role_required("admin")
def admin_database_import():
    db_path = sqlite_main_database_path()
    if not db_path or not db_path.is_file():
        flash("Verilənlər bazasının idxalı yalnız fayl əsaslı SQLite üçün mövcuddur.", "warning")
        return redirect(url_for("admin_settings"))

    user = db.session.get(User, session.get("user_id"))
    password = request.form.get("confirm_password", "")
    if not user or not check_password_hash(user.password_hash, password):
        flash("Şifrə yanlışdır. Idxal ləğv edildi.", "danger")
        return redirect(url_for("admin_settings"))

    upload = request.files.get("backup_file")
    if not upload or not upload.filename:
        flash("Yedək .db faylını seçin.", "danger")
        return redirect(url_for("admin_settings"))

    ext = upload.filename.rsplit(".", 1)[-1].lower() if "." in upload.filename else ""
    if ext not in ("db", "sqlite", "sqlite3"):
        flash("Yalnız SQLite yedək faylı (.db) qəbul edilir.", "warning")
        return redirect(url_for("admin_settings"))

    tmp_fd, tmp_name = tempfile.mkstemp(prefix="emtk_import_", suffix=".db")
    os.close(tmp_fd)
    tmp_path = Path(tmp_name)
    try:
        upload.save(tmp_path)
        ok, err = validate_emtk_sqlite_file(tmp_path)
        if not ok:
            flash(err, "danger")
            return redirect(url_for("admin_settings"))

        stamp = utc_to_local(datetime.now(timezone.utc)).strftime("%Y%m%d_%H%M%S")
        safety_copy = db_path.with_name(f"{db_path.stem}.before_restore_{stamp}{db_path.suffix}")
        try:
            shutil.copy2(db_path, safety_copy)
        except OSError:
            safety_copy = None

        db.session.remove()
        db.engine.dispose()
        try:
            shutil.copy2(tmp_path, db_path)
        except OSError as exc:
            flash(f"Baza faylı əvəz edilmədi: {exc}", "danger")
            return redirect(url_for("admin_settings"))

        try:
            run_startup_migrations()
            get_smtp_config()
        except SQLAlchemyError as exc:
            flash(f"Bərpa sonrası sxem yoxlaması uğursuz: {exc}. Köhnə nüsxə: {safety_copy or 'yoxdur'}.", "danger")
            return redirect(url_for("admin_settings"))

        audit("Verilənlər bazası idxal ilə bərpa edildi (SQLite)")
        session.clear()
        flash(
            "Baza uğurla bərpa edildi. Köhnə fayl təhlükəsizlik nüsxəsi kimi saxlanıla bilər. "
            "Yenidən daxil olun.",
            "success",
        )
        return redirect(url_for("login"))
    finally:
        tmp_path.unlink(missing_ok=True)


@app.route("/admin/reset-financial", methods=["POST"])
@login_required
@role_required("admin")
def admin_reset_financial():
    # Reload from DB to get fresh password_hash (avoids SQLAlchemy identity-map stale data)
    user = db.session.get(User, session.get("user_id"))
    if not user:
        flash("İstifadəçi tapılmadı.", "danger")
        return redirect(url_for("admin_settings"))
    password = request.form.get("confirm_password", "")
    if not check_password_hash(user.password_hash, password):
        flash("Şifrə yanlışdır. Sıfırlama ləğv edildi.", "danger")
        return redirect(url_for("admin_settings"))

    # Delete in dependency order to avoid FK violations
    Payment.query.delete(synchronize_session=False)
    # Remove all invoices (including future months) and expenses.
    Invoice.query.delete(synchronize_session=False)
    Expense.query.delete(synchronize_session=False)
    # Reset apartment credit balances
    for apt in Apartment.query.all():
        apt.credit_balance = Decimal("0.00")
    # Delete standalone financial history
    BalanceTopUp.query.delete(synchronize_session=False)
    AuditLog.query.delete(synchronize_session=False)
    db.session.commit()

    # Write a fresh audit entry after the wipe
    db.session.add(AuditLog(actor_user_id=user.id, action="Bütün maliyyə məlumatları sıfırlandı (admin)"))
    db.session.commit()

    flash("Bütün ödəniş, hesab, xərc, borc və tarixçə məlumatları sıfırlandı.", "success")
    return redirect(url_for("admin_settings"))


@app.route("/admin/health/money-schema")
@login_required
@role_required("admin", "komendant")
def admin_health_money_schema():
    expected = {
        "apartment": {"credit_balance"},
        "tariff": {"amount"},
        "invoice": {"amount", "paid_amount"},
        "payment": {"amount"},
        "expense_template": {"default_amount"},
        "expense": {"amount"},
        "balance_top_up": {"amount"},
    }
    result = {"ok": True, "tables": {}, "dialect": db.engine.dialect.name}
    if not _is_sqlite():
        # Проверка через PRAGMA применима только к SQLite. Для других диалектов
        # полагаемся на Alembic/миграции и возвращаем пустой отчёт.
        result["skipped"] = "non-sqlite dialect"
        return result
    with db.engine.connect() as conn:
        for table_name, money_columns in expected.items():
            pragma_rows = conn.exec_driver_sql(f"PRAGMA table_info({table_name})").fetchall()
            col_type = {row[1]: str(row[2] or "").upper() for row in pragma_rows}
            table_info = {}
            for column in sorted(money_columns):
                actual = col_type.get(column, "")
                is_numeric = "NUMERIC" in actual
                table_info[column] = {"actual_type": actual, "is_numeric": is_numeric}
                if not is_numeric:
                    result["ok"] = False
            result["tables"][table_name] = table_info
    return result


@app.route("/admin/health/calculation-smoke")
@login_required
@role_required("admin", "komendant")
def admin_health_calculation_smoke():
    """
    Run deterministic calculation smoke-checks without DB writes.
    """
    checks = []

    def _check(name: str, actual: float, expected: float):
        actual_rounded = round(float(actual), 2)
        expected_rounded = round(float(expected), 2)
        ok = actual_rounded == expected_rounded
        checks.append(
            {
                "name": name,
                "ok": ok,
                "actual": actual_rounded,
                "expected": expected_rounded,
            }
        )

    # Scenario 1: credit auto-applies to unpaid invoice.
    apt1 = Apartment(area=80.0, credit_balance=Decimal("15.00"))
    inv1 = Invoice(amount=Decimal("100.00"), paid_amount=Decimal("20.00"), status="gozlemede")
    inv1.apartment = apt1
    applied = _apply_credit_to_invoice(inv1)
    _check("credit_applied_amount", applied, 15.00)
    _check("credit_applied_paid_amount", inv1.paid_amount, 35.00)
    _check("credit_applied_credit_left", apt1.credit_balance, 0.00)

    # Scenario 2: payment overpay moves overflow into apartment credit.
    apt2 = Apartment(area=60.0, credit_balance=Decimal("0.00"))
    inv2 = Invoice(amount=Decimal("100.00"), paid_amount=Decimal("95.00"), status="gozlemede")
    inv2.apartment = apt2
    delta_info_overpay = _apply_payment_delta(inv2, 10.00)
    _check("overpay_paid_capped_to_invoice_amount", inv2.paid_amount, 100.00)
    _check("overpay_moved_to_credit", delta_info_overpay["moved_to_credit"], 5.00)
    _check("overpay_credit_balance", apt2.credit_balance, 5.00)

    # Scenario 3: negative correction first consumes apartment credit.
    apt3 = Apartment(area=60.0, credit_balance=Decimal("4.00"))
    inv3 = Invoice(amount=Decimal("100.00"), paid_amount=Decimal("90.00"), status="gozlemede")
    inv3.apartment = apt3
    delta_info_reversal = _apply_payment_delta(inv3, -6.00)
    _check("reversal_removed_from_credit", delta_info_reversal["removed_from_credit"], 4.00)
    _check("reversal_credit_after_consume", apt3.credit_balance, 0.00)
    _check("reversal_paid_after_remainder", inv3.paid_amount, 88.00)

    # Scenario 3b: negative Mədaxil (debt adjustment) lowers paid_amount; balance follows.
    apt3b = Apartment(area=60.0, credit_balance=Decimal("0.00"))
    inv3b = Invoice(amount=Decimal("100.00"), paid_amount=Decimal("49.00"), status="gozlemede")
    inv3b.apartment = apt3b
    _apply_payment_delta(inv3b, -50.00, debt_adjustment=True)
    _check("debt_adjust_paid_amount", inv3b.paid_amount, -1.00)
    _check("debt_adjust_invoice_balance", float(inv3b.paid_amount) - float(inv3b.amount), -101.00)

    # Scenario 4: resident debt uses max(0, base_debt - credit).
    base_debt = max(0.0, 100.0 - 100.0) + max(0.0, 50.0 - 30.0)
    credit_balance = 25.0
    resident_debt = max(0.0, round(base_debt - credit_balance, 2))
    _check("resident_debt_clamped", resident_debt, 0.00)

    # Scenario 5: invoice amount uses tariff scope and types.
    apt4 = Apartment(id=101, area=50.0)
    t1 = Tariff(id=201, type="per_m2", amount=Decimal("1.20"))
    t2 = Tariff(id=202, type="fixed", amount=Decimal("10.00"))
    t3 = Tariff(id=203, type="fixed", amount=Decimal("7.00"))
    scope_map = {203: {999}}  # excluded for this apartment
    amount_total = compute_invoice_amount(apt4, [t1, t2, t3], scope_map)
    _check("invoice_amount_from_tariffs", amount_total, 70.00)

    ok = all(c["ok"] for c in checks)
    return {"ok": ok, "checks": checks}


@app.route("/resident/whatsapp/connect")
@login_required
def resident_whatsapp_connect():
    """Редирект резидента на wa.me/<service_number>. Флаг подключения ставится webhook'ом."""
    wa_cfg = get_whatsapp_config()
    number = _wa_digits(wa_cfg.service_number) if wa_cfg.service_number else None
    if not number:
        flash("WhatsApp servis nömrəsi təyin edilməyib. Administrator ilə əlaqə saxlayın.", "warning")
        return redirect(url_for("dashboard"))
    # Префилл текста в wa.me; привязка пользователя в webhook идёт по номеру телефона из JID.
    user = current_user()
    name = (user.full_name or "").strip() if user else ""
    text = f"Bildirişlərə qoşulmağa icazə verirəm.\n{name}"
    from urllib.parse import quote
    return redirect(f"https://wa.me/{number}?text={quote(text)}")


def _wa_log_webhook(*, status_code: int, event: Optional[str], remote_jid: Optional[str],
                    digits: Optional[str], matched_user_id: Optional[int], note: str,
                    raw_body: str) -> None:
    """Пишет запись в диагностический журнал и чистит старые (>200)."""
    try:
        entry = WhatsappWebhookLog(
            remote_ip=(request.headers.get("X-Forwarded-For") or request.remote_addr or "")[:64],
            status_code=status_code,
            event=(event or "")[:64],
            remote_jid=(remote_jid or "")[:128],
            digits=(digits or "")[:32],
            matched_user_id=matched_user_id,
            note=(note or "")[:255],
            raw_body=(raw_body or "")[:8000],
        )
        db.session.add(entry)
        db.session.commit()
        # Ротация: храним максимум 200 последних записей.
        total = WhatsappWebhookLog.query.count()
        if total > 200:
            old_ids = [
                row.id for row in
                WhatsappWebhookLog.query.order_by(WhatsappWebhookLog.id.asc()).limit(total - 200).all()
            ]
            if old_ids:
                WhatsappWebhookLog.query.filter(WhatsappWebhookLog.id.in_(old_ids)).delete(synchronize_session=False)
                db.session.commit()
    except Exception:
        db.session.rollback()


def _wa_extract_messages(payload: dict) -> list:
    """Вытягивает список сообщений из пейлоада Evolution API в разных возможных формах."""
    if not isinstance(payload, dict):
        return []
    data = payload.get("data")
    if data is None:
        # Иногда корневой объект — это и есть message.
        return [payload] if payload.get("key") else []
    if isinstance(data, list):
        return [m for m in data if isinstance(m, dict)]
    if isinstance(data, dict):
        if isinstance(data.get("messages"), list):
            return [m for m in data["messages"] if isinstance(m, dict)]
        return [data]
    return []


def _wa_is_from_me(key: dict) -> bool:
    """fromMe может прийти как bool/str/int."""
    v = key.get("fromMe")
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    if isinstance(v, str):
        return v.strip().lower() in ("1", "true", "yes")
    return False


@csrf.exempt
@app.route("/whatsapp/webhook", methods=["POST"])
@app.route("/whatsapp/webhook/<path:subpath>", methods=["POST"])
def whatsapp_webhook(subpath: Optional[str] = None):
    """
    Webhook от Evolution API. Принимает как единый URL (webhookByEvents=false),
    так и URL с суффиксами (/messages-upsert и т.п., если включено webhookByEvents=true).

    Секрет передаётся в query (?secret=...) или заголовке X-Webhook-Secret / apikey.
    """
    raw_body = request.get_data(as_text=True) or ""

    wa_cfg = get_whatsapp_config()
    expected = (wa_cfg.webhook_secret or "").strip()
    provided = (
        request.args.get("secret")
        or request.headers.get("X-Webhook-Secret")
        or request.headers.get("Apikey")
        or request.headers.get("apikey")
        or ""
    )
    if not expected or not hmac.compare_digest(provided, expected):
        _wa_log_webhook(
            status_code=403, event=None, remote_jid=None, digits=None,
            matched_user_id=None, note="forbidden: secret mismatch",
            raw_body=raw_body,
        )
        return ("forbidden", 403)

    try:
        payload = request.get_json(force=True, silent=True) or {}
    except Exception:
        payload = {}

    event = (payload.get("event") or payload.get("type") or "").lower()
    # Если включён webhookByEvents=true, Evolution добавляет суффикс (напр. /messages-upsert)
    # и событие дублируется в URL — учитываем и это.
    if subpath and not event:
        event = subpath.lower().replace("-", ".").replace("_", ".")

    messages = _wa_extract_messages(payload)

    if not messages:
        _wa_log_webhook(
            status_code=200, event=event or None, remote_jid=None, digits=None,
            matched_user_id=None, note=f"received ok, no messages in payload (event={event or 'n/a'})",
            raw_body=raw_body,
        )
        try:
            wa_queue_drain_once()
        except Exception:
            pass
        return ("ok", 200)

    updated = 0
    for m in messages:
        key = m.get("key") or {}
        from_me = _wa_is_from_me(key)
        jid = key.get("remoteJid") or m.get("remoteJid") or ""
        # sender в группах может быть в participant
        participant = key.get("participant") or m.get("participant") or ""
        effective_jid = participant or jid  # для личных чатов participant пуст — берём remoteJid

        if from_me:
            _wa_log_webhook(
                status_code=200, event=event or None, remote_jid=jid, digits=None,
                matched_user_id=None, note="skipped: fromMe=true",
                raw_body=raw_body,
            )
            continue
        if not effective_jid:
            _wa_log_webhook(
                status_code=200, event=event or None, remote_jid=None, digits=None,
                matched_user_id=None, note="skipped: no remoteJid",
                raw_body=raw_body,
            )
            continue

        jid_left = effective_jid.split("@", 1)[0]
        jid_suffix = effective_jid.split("@", 1)[1] if "@" in effective_jid else ""
        digits = re.sub(r"\D+", "", jid_left)

        # @lid — внутренний идентификатор, не телефон. Матч всё равно попробуем, но в note отметим.
        lid_warning = "; jid=@lid (внутренний ID, не телефон)" if jid_suffix.lower() == "lid" else ""

        if not digits:
            _wa_log_webhook(
                status_code=200, event=event or None, remote_jid=effective_jid,
                digits=None, matched_user_id=None,
                note=f"skipped: cannot extract digits{lid_warning}",
                raw_body=raw_body,
            )
            continue

        candidates = User.query.filter(User.phone.isnot(None)).all()
        user = None
        for u in candidates:
            u_digits = re.sub(r"\D+", "", u.phone or "")
            if u_digits and (u_digits == digits or u_digits.endswith(digits) or digits.endswith(u_digits)):
                user = u
                break

        if not user:
            _wa_log_webhook(
                status_code=200, event=event or None, remote_jid=effective_jid,
                digits=digits, matched_user_id=None,
                note=f"no user match{lid_warning}",
                raw_body=raw_body,
            )
            continue

        user.whatsapp_connected = True
        user.whatsapp_jid = effective_jid[:64]
        user.whatsapp_connected_at = datetime.now(timezone.utc)
        updated += 1
        _wa_log_webhook(
            status_code=200, event=event or None, remote_jid=effective_jid,
            digits=digits, matched_user_id=user.id,
            note=f"matched user #{user.id} ({user.full_name}){lid_warning}",
            raw_body=raw_body,
        )

    if updated:
        db.session.commit()

    try:
        wa_queue_drain_once()
    except Exception:
        pass

    return ("ok", 200)


@app.route("/admin/whatsapp/logs")
@login_required
@role_required("admin", "komendant")
def admin_whatsapp_logs():
    """Админский просмотр последних webhook-событий."""
    logs = WhatsappWebhookLog.query.order_by(WhatsappWebhookLog.id.desc()).limit(100).all()
    return render_template("admin_whatsapp_logs.html", logs=logs)


if __name__ == "__main__":
    with app.app_context():
        run_startup_migrations()
        get_smtp_config()
    host = os.getenv("FLASK_HOST", "0.0.0.0")
    port = int(os.getenv("FLASK_PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG", "0") == "1"
    app.run(host=host, port=port, debug=debug)
